import os
import glob
import csv
import numpy as np
import pandas as pd
from scipy import stats
import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def find_best_iteration(filepath):
    """First pass to find best iteration based on validation data"""
    with open(filepath, 'r') as file:
        content = file.read()

    parts = content.split('--- ITERATION RESULTS ---')
    second_table = pd.read_csv(io.StringIO(parts[1].strip()))
    
    # Filter for validation rows
    validation_rows = second_table[second_table['label'] == 'validation']
    
    if validation_rows.empty:
        # If no validation data, fall back to all data
        print(f"Warning: No validation data found in {filepath}. Using all data.")
        validation_rows = second_table
    
    # Find the maximum F1 score
    max_f1 = validation_rows['f1'].max()
    
    # Filter rows with maximum F1 score
    max_f1_rows = validation_rows[validation_rows['f1'] == max_f1]
    
    # Among these rows, find the one with minimum score
    best_validation_row = max_f1_rows.loc[max_f1_rows['score'].idxmin()]
    
    return best_validation_row['iteration']


def process_file(filepath):
    """
    Process a file using a two-pass approach:
    1. Find the best iteration based on validation data
    2. Get test metrics at the best iteration
    """
    #print(f"Processing file: {filepath}")

    # First pass: find the best iteration based on validation data
    best_iteration = find_best_iteration(filepath)
    #print(f"Best iteration found: {best_iteration}")
    
    # Second pass: get test metrics at the best iteration
    with open(filepath, 'r') as file:
        content = file.read()

    parts = content.split('--- ITERATION RESULTS ---')
    second_table = pd.read_csv(io.StringIO(parts[1].strip()))
    
    # Filter for test rows
    test_rows = second_table[second_table['label'] == 'test'].copy()
    
    if test_rows.empty:
        # If no test data, fall back to all data
        print(f"Warning: No test data found in {filepath}. Using all data.")
        test_rows = second_table.copy()
    
    # Find the closest iteration to the best one
    test_rows.loc[:, 'iteration_diff'] = abs(test_rows['iteration'] - best_iteration)
    closest_test_row = test_rows.loc[test_rows['iteration_diff'].idxmin()]
    
    #print(f"Using test metrics from iteration {closest_test_row['iteration']}")
    #print(f"Selected test row:\n{closest_test_row[['iteration', 'score', 'f1', 'smoothness', 'avg_freeze']]}")

    result = {
        'score': closest_test_row['score'],
        'f1': closest_test_row['f1'],
        'smoothness': closest_test_row['smoothness'],
        'freeze': closest_test_row['avg_freeze']
    }

    #print(f"Returned result: {result}")
    return result


def get_model_results(path):
    results = {}
    for file in glob.glob(os.path.join(path, '*.csv')):
        sim_num = int(file.split('_')[-1].split('.')[0])
        results[sim_num] = process_file(file)
    return results


def process_directory(path):
    """
    Process a directory containing CSV result files.
    
    Parameters
    ----------
    path : str
        Path to the directory containing CSV files
    
    Returns
    -------
    dict
        Dictionary with simulation numbers as keys and results as values
    """
    results = {}
    csv_files = glob.glob(os.path.join(path, '*.csv'))
    
    if not csv_files:
        print(f"No CSV files found in {path}")
        return {}
    
    print(f"Found {len(csv_files)} CSV files in {path}")
    
    for file in csv_files:
        try:
            # Extract simulation number from filename
            filename = os.path.basename(file)
            if '_' in filename:
                sim_num = int(filename.split('_')[-1].split('.')[0])
            else:
                sim_num = int(filename.split('.')[0])
                
            # Process the file
            results[sim_num] = process_file(file)
            print(f"  Processed {filename}, simulation {sim_num}")
        except Exception as e:
            print(f"  Error processing {file}: {str(e)}")
    
    return results


def apply_fdr_correction(p_values):
    """
    Apply Benjamini-Hochberg procedure for controlling false discovery rate (FDR)
    
    Parameters:
    -----------
    p_values : list
        List of p-values to correct
        
    Returns:
    --------
    list
        List of corrected p-values in the same order as the input
    """
    if not p_values:
        return []
    
    # Create a list of (index, p-value) tuples
    indexed_p_values = list(enumerate(p_values))
    
    # Sort by p-value
    indexed_p_values.sort(key=lambda x: x[1])
    
    # Get the total number of tests
    m = len(p_values)
    
    # Calculate the critical values and create new p-values
    corrected_p_values = [None] * m
    
    # Apply BH procedure
    for rank, (original_index, p_value) in enumerate(indexed_p_values, 1):
        # Calculate BH critical value
        critical_value = (rank / m) * 0.05
        
        # The corrected p-value is p * m / rank
        corrected_p = min(p_value * m / rank, 1.0)
        
        # Store in the order of the original indices
        corrected_p_values[original_index] = corrected_p
    
    # Ensure monotonicity (each corrected p-value >= previous one when sorted by original p-value)
    for i in range(len(indexed_p_values) - 2, -1, -1):
        original_index = indexed_p_values[i][0]
        next_original_index = indexed_p_values[i + 1][0]
        
        if corrected_p_values[original_index] > corrected_p_values[next_original_index]:
            corrected_p_values[next_original_index] = corrected_p_values[original_index]
    
    return corrected_p_values


def perform_statistical_test(control_data, test_data, test_type='ttest'):
    if test_type == 'ttest':
        statistic, p_value = stats.ttest_rel(control_data, test_data)
    elif test_type == 'wilcoxon':
        statistic, p_value = stats.wilcoxon(control_data, test_data)
    else:
        raise ValueError("Invalid test type. Choose 'ttest' or 'wilcoxon'.")

    return statistic, p_value


def run_mccv_analysis(control_path, test_paths, test_type='ttest'):
    """
    Run Monte Carlo Cross-Validation analysis on control and test directories.

    Parameters:
    -----------
    control_path : str
        Path to the control directory with CSV files
        Special value "BEST_PER_SCORE" means use best model per metric as reference
    test_paths : list
        List of paths to test directories with CSV files
    test_type : str
        Statistical test to perform ('ttest' or 'wilcoxon')
        
    Returns:
    --------
    dict
        Dictionary of results with metrics as keys
    """
    # Special case: Best per score mode
    best_per_score_mode = (control_path == "BEST_PER_SCORE")
    
    if best_per_score_mode:
        print("Using best model per metric as the reference")
        all_paths = test_paths
    else:
        # Standard mode - use the specified control path
        all_paths = [control_path] + test_paths
    
    # Collect all model results
    all_model_results = {}
    for path in all_paths:
        model_name = os.path.basename(path)
        print(f"\nProcessing {model_name}...")
        all_model_results[model_name] = get_model_results(path)
    
    # Initialize the results dictionary
    results = {'f1': {}, 'score': {}, 'freeze': {}, 'smoothness': {}}
    
    # Process each metric separately for best per score mode
    if best_per_score_mode:
        for metric in ['f1', 'score', 'freeze', 'smoothness']:
            # First, calculate mean and standard deviation for each model
            metric_stats = {}
            for model_name, model_data in all_model_results.items():
                # Extract values for this metric from all simulations
                values = [sim_data[metric] for sim_num, sim_data in model_data.items() 
                         if metric in sim_data]
                
                if values:
                    metric_stats[model_name] = {
                        'mean': np.mean(values),
                        'std': np.std(values),
                        'values': values
                    }
            
            # Skip if no models have this metric
            if not metric_stats:
                continue
                
            # Find the best model for this metric
            if metric == 'score':  # Lower is better
                best_model = min(metric_stats.keys(), 
                                key=lambda m: metric_stats[m]['mean'])
                print(f"\n===== {metric.upper()} =====")
                print(f"Finding best model for {metric}...")
                print(f"Best model: {best_model} with {metric} = {metric_stats[best_model]['mean']} (lower is better)")
            else:  # Closer to 1 is better
                best_model = min(metric_stats.keys(), 
                               key=lambda m: abs(metric_stats[m]['mean'] - 1))
                print(f"\n===== {metric.upper()} =====")
                print(f"Finding best model for {metric}...")
                print(f"Best model: {best_model} with {metric} = {metric_stats[best_model]['mean']} (closer to 1 is better)")
            
            # Add best model to results
            results[metric][best_model] = {
                'mean': metric_stats[best_model]['mean'],
                'std': metric_stats[best_model]['std']
            }
            
            # Test each model against the best model
            print(f"Testing all models against reference: {best_model}")
            best_model_values = metric_stats[best_model]['values']
            
            # Store uncorrected p-values temporarily
            uncorrected_results = {}
            p_values = []
            model_names = []
            
            for model_name, model_stats in metric_stats.items():
                if model_name == best_model:
                    continue  # Skip comparing the best model to itself
                
                # Get values for the current model
                model_values = model_stats['values']
                
                # Ensure common number of values
                common_len = min(len(best_model_values), len(model_values))
                if common_len < 2:
                    print(f"Skipping {model_name}: Not enough common data points with {best_model}")
                    continue
                
                # Perform statistical test
                statistic, p_value = perform_statistical_test(
                    best_model_values[:common_len], 
                    model_values[:common_len], 
                    test_type
                )
                
                # Store results temporarily
                uncorrected_results[model_name] = {
                    'mean': model_stats['mean'],
                    'std': model_stats['std'],
                    'difference': model_stats['mean'] - metric_stats[best_model]['mean'],
                    'statistic': statistic,
                    'p_value': p_value
                }
                
                # Save p-value for correction
                p_values.append(p_value)
                model_names.append(model_name)
            
            # Apply FDR correction to all p-values for this metric
            if p_values:
                print(f"  Applying FDR correction to {len(p_values)} p-values...")
                corrected_p_values = apply_fdr_correction(p_values)
                
                # Update results with corrected p-values
                for i, model_name in enumerate(model_names):
                    # Store original p-value for reference
                    uncorrected_p = uncorrected_results[model_name]['p_value']
                    corrected_p = corrected_p_values[i]
                    
                    # Update the result with corrected p-value
                    results[metric][model_name] = {
                        'mean': uncorrected_results[model_name]['mean'],
                        'std': uncorrected_results[model_name]['std'],
                        'difference': uncorrected_results[model_name]['difference'],
                        'statistic': uncorrected_results[model_name]['statistic'],
                        'p_value': corrected_p,
                        'p_value_uncorrected': uncorrected_p
                    }
                    
                    is_significant = corrected_p < 0.05
                    print(f"    {metric} p-value: {uncorrected_p:.4f} (corrected: {corrected_p:.4f}) {'(significant)' if is_significant else ''}")
                
    else:
        # Standard mode (one specific control model)
        # Extract control model results
        control_name = os.path.basename(control_path)
        control_results = all_model_results[control_name]
        
        # Store control model metrics in results
        for metric in ['f1', 'score', 'freeze', 'smoothness']:
            # Extract control values for this metric
            control_values = [data[metric] for sim, data in control_results.items() 
                            if metric in data]
            
            if control_values:
                results[metric][control_name] = {
                    'mean': np.mean(control_values),
                    'std': np.std(control_values)
                }
        
        # For each metric, we will collect all p-values before applying FDR correction
        for metric in ['f1', 'score', 'freeze', 'smoothness']:
            uncorrected_results = {}
            p_values = []
            model_names = []

            # Process test models
            for model_path in test_paths:
                model_name = os.path.basename(model_path)
                test_results = all_model_results[model_name]
                
                # Extract values for this metric from both control and test results
                control_data = []
                test_data = []
                
                # Use simulation numbers that are common to both models
                common_sims = set(control_results.keys()) & set(test_results.keys())

                # Skip metrics that aren't present in both models
                if not all(metric in control_results.get(sim, {}) and 
                          metric in test_results.get(sim, {})
                          for sim in common_sims):
                    print(f"Skipping {metric} for {model_name}: not present in all simulations")
                    continue
                
                # Collect data from common simulations
                for sim in common_sims:
                    if metric in control_results[sim] and metric in test_results[sim]:
                        control_data.append(control_results[sim][metric])
                        test_data.append(test_results[sim][metric])
                
                # Skip if not enough data
                if len(control_data) < 2 or len(test_data) < 2:
                    print(f"Skipping {metric} for {model_name}: not enough data points")
                    continue
                
                # Calculate means
                test_mean = np.mean(test_data)
                control_mean = np.mean(control_data)
                
                # Perform statistical test
                statistic, p_value = perform_statistical_test(control_data, test_data, test_type)

                # Store results temporarily
                uncorrected_results[model_name] = {
                    'mean': test_mean,
                    'std': np.std(test_data),
                    'difference': test_mean - control_mean,
                    'statistic': statistic,
                    'p_value': p_value
                }
                
                # Save p-value for correction
                p_values.append(p_value)
                model_names.append(model_name)
            
            # Apply FDR correction to all p-values for this metric
            if p_values:
                print(f"Applying FDR correction to {len(p_values)} p-values for {metric}...")
                corrected_p_values = apply_fdr_correction(p_values)
                
                # Update results with corrected p-values
                for i, model_name in enumerate(model_names):
                    # Store original p-value for reference
                    uncorrected_p = uncorrected_results[model_name]['p_value']
                    corrected_p = corrected_p_values[i]
                    
                    # Update the result with corrected p-value
                    results[metric][model_name] = {
                        'mean': uncorrected_results[model_name]['mean'],
                        'std': uncorrected_results[model_name]['std'],
                        'difference': uncorrected_results[model_name]['difference'],
                        'statistic': uncorrected_results[model_name]['statistic'],
                        'p_value': corrected_p,
                        'p_value_uncorrected': uncorrected_p
                    }
                    
                    is_significant = corrected_p < 0.05
                    print(f"    {metric} p-value: {uncorrected_p:.4f} (corrected: {corrected_p:.4f}) {'(significant)' if is_significant else ''}")

    return results


def save_results_to_csv(results, output_dir):
    """
    Save the analysis results to an Excel file with formatting.
    - Values are rounded to n decimal places
    - Significant values (p < 0.05) are marked with an asterisk (*)
    - Best performing metrics in each column are bolded:
      - For F1, LDJ, and Dampening: values closest to 1 are best
      - For Distance: lowest values are best
    - When best values are not significantly different from control, all tied values are bolded
    """
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    num_dec = 3
    # Function to parse the model name into parameter dictionary
    def parse_model_params(model_name):
        # Start with an empty dictionary for the parameters
        params = {}
        param_order = []  # Track the order of parameters
        
        # Remove control suffix if present, before parsing parameters
        if model_name.endswith(';control'):
            model_name = model_name[:-8]  # Remove ';control' suffix
        
        # Return empty params if model name does not contain parameters
        if '-' not in model_name:
            return params, param_order
            
        # Split by semicolon to get parameter pairs
        param_pairs = model_name.split(';')
        
        for pair in param_pairs:
            if '-' in pair:
                # Split by hyphen to get parameter category and value
                parts = pair.split('-', 1)  # Split on first hyphen only
                if len(parts) == 2:
                    category, value = parts
                    
                    # Format the parameter category (replace _ with space, capitalize words)
                    # Using a standardized way to format parameter names to avoid duplication
                    formatted_category = ' '.join(word.capitalize() for word in category.split('_'))
                    
                    # Ensure consistent naming for Geo Params
                    if formatted_category.lower() == "geo params":
                        formatted_category = "Geo Params"
                    
                    # Format the parameter value if it's a string (contains alphabetic characters)
                    if any(c.isalpha() for c in value):
                        value = ' '.join(word.capitalize() for word in value.split('_'))
                    
                    params[formatted_category] = value
                    param_order.append(formatted_category)
        
        # Handle special cases for geometry parameters with implicit values
        geometry_param_mapping = {
            'Ellipse': '2',
            'Toroid': '3',
            'Linear': '3',
            'Klein Bottle': '3',
            'Mobius Strip': '2',
            'Spiral': '2',
            'Cocentric Circles': '2',
            'Torus': '3',
            'Sphere': '3',
        }
        
        # Find the geometry and geo params keys if they exist (case-insensitive)
        geometry_key = next((key for key in params.keys() if key.lower() == "geometry"), None)
        geo_params_key = next((key for key in params.keys() if key.lower() == "geo params"), None)
        
        # If there's a Geometry parameter and either:
        # 1. No Geo Params key exists, or
        # 2. Geo Params key exists but has a blank value
        if geometry_key:
            geo_value = params[geometry_key]
            
            # Check if geo params is missing or has a blank value
            missing_geo_params = (geo_params_key is None or 
                                  not params[geo_params_key] or 
                                  params[geo_params_key].strip() == 'Na' or
                                  params[geo_params_key].strip() == '')
            
            if missing_geo_params:
                # Look for a matching geometry to assign implicit value
                for geometry_name, implicit_value in geometry_param_mapping.items():
                    if geometry_name.lower() in geo_value.lower():
                        # If geo params key already exists, update its value
                        if geo_params_key:
                            params[geo_params_key] = implicit_value
                        else:
                            # Otherwise create a new entry with standard formatting
                            params["Geo Params"] = implicit_value
                            # Add Geo Params after Geometry in the parameter order
                            if geometry_key in param_order:
                                geo_idx = param_order.index(geometry_key)
                                param_order.insert(geo_idx + 1, "Geo Params")
                            else:
                                param_order.append("Geo Params")
                        break
        
        return params, param_order

    # Maps metrics to their new column names
    metric_column_names = {
        'f1': 'F1',
        'score': 'Distance',
        'freeze': 'Dampening',
        'smoothness': 'LDJ'
    }
    
    # Get all unique models across all metrics
    all_models = set()
    for metric in results.keys():
        all_models.update(results[metric].keys())
    
    # Extract parameters from all models once
    all_param_categories = set()
    model_params = {}
    param_orders = {}
    all_param_orders = []  # Collect all parameter orders to determine a consistent order
    
    for model in all_models:
        model_basename = os.path.basename(model)
        params, param_order = parse_model_params(model_basename)
        model_params[model] = params
        param_orders[model] = param_order
        all_param_categories.update(params.keys())
        all_param_orders.append(param_order)
    
    # Create a consistent order for all parameters based on their relative positions
    sorted_models = sorted(all_param_orders, key=len, reverse=True)
    reference_order = sorted_models[0] if sorted_models else []
    
    # Add any missing parameters that exist in other models
    for order in all_param_orders:
        for param in order:
            if param not in reference_order:
                reference_order.append(param)
    
    # Use this reference order for all parameter columns
    param_columns = reference_order
    
    # Prepare rows for the combined table
    combined_rows = []
    
    # Check if we're in "best per score" mode - this mode has different control models per metric
    best_per_score_mode = False
    
    # In best_per_score mode, every metric has its own "best" model as control
    # We can detect this by checking if there's no common control across metrics
    control_models = set()
    for metric in results:
        if metric in ['f1', 'score', 'freeze', 'smoothness'] and results[metric]:
            # Find the model that doesn't have 'difference', 'statistic', or 'p_value' keys
            # This is the control/reference model for this metric
            for model, values in results[metric].items():
                if not any(k in values for k in ['difference', 'statistic', 'p_value']):
                    control_models.add(model)
                    break
    
    # If we have multiple control models, we're in best_per_score mode
    best_per_score_mode = len(control_models) > 1
    
    if best_per_score_mode:
        print("Detected 'best per score' mode - each metric has its own best model as reference")
        # In this mode, we don't have a global control_name, each metric has its own
        control_name = None
    else:
        # Standard mode - use the first model in the first metric as control
        control_name = list(results[list(results.keys())[0]].keys())[0]  # Assume the first model in the first metric is the control
    
    # Process each model
    for model in all_models:
        # In best_per_score mode, a model can be control for some metrics but not others
        # In standard mode, is_control is True only for the global control model
        is_control = (not best_per_score_mode) and (model == control_name)
        
        # Initialize row with model and parameters
        row = {
            'Model': os.path.basename(model)
        }
        
        # Add parameter values
        for param in param_columns:
            row[param] = model_params[model].get(param, 'NA')
        
        # Add metric values
        for metric in ['f1', 'score', 'freeze', 'smoothness']:
            if metric in results and model in results[metric]:
                values = results[metric][model]
                
                # Round mean and std to 2 decimal places
                row[metric_column_names[metric]] = round(values['mean'], num_dec)
                row[f"{metric_column_names[metric]} Std"] = round(values['std'], num_dec)
                
                # Determine if this model is the control for this specific metric
                metric_control = False
                if best_per_score_mode:
                    # In best_per_score mode, the control doesn't have difference/statistic/p_value
                    metric_control = not any(k in values for k in ['difference', 'statistic', 'p_value'])
                
                # Add difference, statistic, p-value for non-control models
                # (either not global control in standard mode or not metric control in best_per_score mode)
                if (not is_control and not metric_control) and 'difference' in values:
                    row[f"{metric_column_names[metric]} Diff"] = round(values['difference'], num_dec)
                    row[f"{metric_column_names[metric]} Stat"] = round(values['statistic'], num_dec)
                    # Use the corrected p-value if available
                    if 'p_value_uncorrected' in values:
                        row[f"{metric_column_names[metric]} p"] = values['p_value']  # This is the corrected p-value
                        row[f"{metric_column_names[metric]} p_uncorr"] = values['p_value_uncorrected']
                    else:
                        row[f"{metric_column_names[metric]} p"] = values['p_value']
                else:
                    # This model is a control for this metric
                    row[f"{metric_column_names[metric]} Diff"] = 'NA'
                    row[f"{metric_column_names[metric]} Stat"] = 'NA'
                    row[f"{metric_column_names[metric]} p"] = 'NA'
                    if 'p_value_uncorrected' in next(iter(results.get(metric, {}).values())) if results.get(metric, {}) else {}:
                        row[f"{metric_column_names[metric]} p_uncorr"] = 'NA'
            else:
                # If this metric doesn't have data for this model, use NA
                row[metric_column_names[metric]] = 'NA'
                row[f"{metric_column_names[metric]} Std"] = 'NA'
                if not is_control:
                    row[f"{metric_column_names[metric]} Diff"] = 'NA'
                    row[f"{metric_column_names[metric]} Stat"] = 'NA'
                    row[f"{metric_column_names[metric]} p"] = 'NA'
                    if 'p_value_uncorrected' in next(iter(results.get(metric, {}).values())) if results.get(metric, {}) else {}:
                        row[f"{metric_column_names[metric]} p_uncorr"] = 'NA'
        
        combined_rows.append(row)
    
    # Create DataFrame from combined rows
    df = pd.DataFrame(combined_rows)
    
    # Add the calculated column: Distance*Dampening/F1
    df['Distance*Dampening/F1'] = df.apply(
        lambda row: (
            float(row['Distance']) * float(row['Dampening']) / float(row['F1'])
            if (row['Distance'] != 'NA' and row['Dampening'] != 'NA' and row['F1'] != 'NA' and float(row['F1']) != 0)
            else 'NA'
        ),
        axis=1
    )
    
    # Sort the DataFrame by the calculated column (Distance*Dampening/F1)
    # First convert the column to numeric, with errors='coerce' to handle 'NA' values
    df['Sort_Value'] = pd.to_numeric(df['Distance*Dampening/F1'], errors='coerce')
    # Sort by the numeric column, with NA values at the end
    df = df.sort_values(by='Sort_Value', ascending=True, na_position='last')
    # Drop the temporary sorting column
    df = df.drop(columns=['Sort_Value'])
    
    # Define column order
    base_columns = ['Model'] + param_columns
    metric_columns = []
    for metric in ['F1', 'Distance', 'Dampening', 'LDJ']:
        metric_columns.append(metric)
        metric_columns.append(f"{metric} Std")
        # Only add difference columns for the standard deviation, statistics and p-values
        # if there are non-control models
        if len(all_models) > 1:
            metric_columns.append(f"{metric} Diff")
            metric_columns.append(f"{metric} Stat")
            metric_columns.append(f"{metric} p")
            if 'p_value_uncorrected' in next(iter(results.get('f1', {}).values())) if results.get('f1', {}) else {}:
                metric_columns.append(f"{metric} p_uncorr")
    
    # Add the calculated column at the end
    final_columns = base_columns + metric_columns + ['Distance*Dampening/F1']
    
    # Reorder columns and only include columns that exist
    existing_columns = [col for col in final_columns if col in df.columns]
    df = df[existing_columns]
    
    # Format p-values and apply asterisks to significant values
    p_value_columns = [col for col in df.columns if ' p' in col and 'p_uncorr' not in col]
    for col in p_value_columns:
        metric_col = col.split(' p')[0]
        # Create a new column with asterisks for significant values
        df[metric_col] = df.apply(
            lambda row: str(row[metric_col]) + "*" if pd.notnull(row[col]) and row[col] != 'NA' and float(row[col]) < 0.05 else str(row[metric_col]),
            axis=1
        )
    
    # Save to Excel file
    file_name = f"combined_mccv_analysis_{timestamp}.xlsx"
    file_path = os.path.join(output_dir, file_name)

    # Need to import these libraries for Excel formatting
    try:
        import openpyxl
        from openpyxl.styles import Font
        
        # Create Excel writer with openpyxl engine
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Results')
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Results']
            
            # Format headers - make them bold with light gray background
            header_font = Font(bold=True)
            for col_idx, col_name in enumerate(df.columns):
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                cell = worksheet[f"{col_letter}1"]
                cell.font = header_font
                
                # Hide standard deviation, p-value, statistics, and difference columns
                if any(suffix in col_name for suffix in [' Std', ' p', ' Stat', ' Diff']):
                    worksheet.column_dimensions[col_letter].hidden = True
            
            # Set column width for visible columns
            for col_idx, col_name in enumerate(df.columns):
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                # Only adjust width for visible columns
                if not any(suffix in col_name for suffix in [' Std', ' p', ' Stat', ' Diff']):
                    max_length = 0
                    for row_idx in range(1, len(df) + 2):  # Include header and all rows
                        cell = worksheet[f"{col_letter}{row_idx}"]
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = max(max_length + 2, 10)  # Add padding and minimum width
                    worksheet.column_dimensions[col_letter].width = adjusted_width
            
            # Set the first column a bit wider for the model names
            if 'A' in worksheet.column_dimensions:
                worksheet.column_dimensions['A'].width = 30
            
            # Bold the best metrics in each column
            metric_names = ['F1', 'Distance', 'Dampening', 'LDJ']
            
            for metric in metric_names:
                # Get index of the metric column
                if metric not in df.columns:
                    continue
                    
                col_idx = df.columns.get_loc(metric)
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                
                # Get all values for this metric
                metric_values = []
                for row_idx, value in enumerate(df[metric]):
                    # Skip NA values
                    if value == 'NA' or pd.isna(value):
                        continue
                        
                    # Check if the value has an asterisk (indicating significance)
                    has_asterisk = False
                    numeric_value = value
                    if isinstance(value, str):
                        if value.endswith('*'):
                            has_asterisk = True
                            numeric_value = value[:-1]  # Remove asterisk for numeric comparison
                    
                    try:
                        # Store as (row_index, model_name, numeric_value, has_asterisk)
                        metric_values.append((row_idx, df.iloc[row_idx]['Model'], float(numeric_value), has_asterisk))
                    except (ValueError, TypeError):
                        continue
                
                if not metric_values:
                    continue
                
                # Determine best value based on metric type (higher or lower is better)
                if metric == 'Distance':
                    # For distance, lower is better
                    best_value = min(value for _, _, value, _ in metric_values)
                    best_indices = [idx for idx, _, value, _ in metric_values if value == best_value]
                else:
                    # For F1, Dampening, LDJ, closer to 1 is better
                    closest_to_one = min(abs(value - 1) for _, _, value, _ in metric_values)
                    best_indices = [idx for idx, _, value, _ in metric_values if abs(value - 1) == closest_to_one]
                
                # Initialize list for indices to bold - just the best performers initially
                bold_indices = best_indices.copy()
                
                # Get the best model for this metric
                if best_indices:
                    best_idx = best_indices[0]
                    best_model = df.iloc[best_idx]['Model']
                else:
                    best_model = None
                
                # Get p-value column name for this metric
                p_column = f"{metric} p"
                
                if best_per_score_mode:
                    # In "best per score" mode, bold the best value and any not significantly different from it
                    if p_column in df.columns and best_model:
                        for idx, model, value, _ in metric_values:
                            if idx in best_indices:
                                continue  # Already in bold_indices
                            
                            # Check if this model is significantly different from the best
                            row_p_value = df.iloc[idx].get(p_column, None)
                            if row_p_value is not None and row_p_value != 'NA':
                                try:
                                    # If p >= 0.05, this model is not significantly different from the best
                                    if float(row_p_value) >= 0.05:
                                        bold_indices.append(idx)
                                        print(f"Bolding {model} for {metric} (not significantly different from best)")
                                except (ValueError, TypeError) as e:
                                    print(f"Error parsing p-value '{row_p_value}': {str(e)}")
                else:
                    # Standard mode - find control and check significance against it
                    # Find control index and value
                    control_idx = None
                    control_value = None
                    for idx, model, value, _ in metric_values:
                        if model == os.path.basename(control_name):
                            control_idx = idx
                            control_value = value
                            break
                    
                    # Check if the best performer is significantly different from the control
                    # If not, bold both the best and the control
                    if control_idx is not None and control_idx not in best_indices:
                        # First, check if the best performer is significantly different from control
                        best_significantly_different = False
                        
                        # Look through all the best models
                        for idx in best_indices:
                            row_data = df.iloc[idx]
                            if p_column in row_data and row_data[p_column] != 'NA':
                                try:
                                    # If p < 0.05, the best is significantly different from control
                                    if float(row_data[p_column]) < 0.05:
                                        best_significantly_different = True
                                        break
                                except (ValueError, TypeError) as e:
                                    print(f"Error parsing p-value: {str(e)}")
                        
                        # If best model is NOT significantly different from control, bold the control too
                        # This is the ONLY case where we bold something other than the best performer
                        if not best_significantly_different:
                            bold_indices.append(control_idx)
                
                # Apply bold formatting
                for idx in bold_indices:
                    cell = worksheet[f"{col_letter}{idx + 2}"]  # +2 for header row and 1-indexing
                    cell.font = Font(bold=True)
     
        print(f"Combined results saved to {file_path} with formatting")
        print(f"Statistical columns are hidden in the Excel output")
        print(f"Significant values are marked with asterisks (*)")
        
        if best_per_score_mode:
            print(f"Best per score mode: Best value and all statistically equivalent values are bolded in each column")
        else:
            print(f"Best metrics in each column are bolded")
        
        '''# Generate LaTeX table
        try:
            latex_file = export_to_latex(
                file_path, 
                caption=f"MCCV Analysis Results ({timestamp})",
                label=f"tab:mccv-analysis-{timestamp}"
            )
            print(f"LaTeX table saved to {latex_file}")
        except Exception as e:
            print(f"Warning: Failed to generate LaTeX table: {str(e)}")'''
        
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        print("Warning: openpyxl not available. Saving as CSV without formatting.")
        csv_file_path = os.path.join(output_dir, f"combined_mccv_analysis_{timestamp}.csv")
        df.to_csv(csv_file_path, index=False)
        print(f"Combined results saved to {csv_file_path}")

# Debugging function to print raw data
def print_raw_data(results):
    """Print raw data from analysis results for debugging"""
    print("\n=== RAW DATA ===")
    for metric, models in results.items():
        print(f"\nMetric: {metric}")
        for model_name, values in models.items():
            print(f"  Model: {model_name}")
            for key, value in values.items():
                # Format numeric values for readability
                if isinstance(value, (int, float)):
                    print(f"    {key}: {value:.4f}")
                else:
                    print(f"    {key}: {value}")
            print("  ---")


def export_to_latex(excel_file, output_file=None, caption=None, label=None):
    """
    Export the Excel results file to LaTeX format.
    
    Parameters:
    -----------
    excel_file : str
        Path to the Excel file to convert
    output_file : str, optional
        Path where to save the LaTeX file. If None, a file with the same
        name but .tex extension will be created in the same directory.
    caption : str, optional
        Caption for the LaTeX table
    label : str, optional
        Label for the LaTeX table (for cross-referencing)
        
    Returns:
    --------
    str
        Path to the generated LaTeX file
    """
    try:
        from excel2latex import excel_to_latex
    except ImportError:
        from .excel2latex import excel_to_latex
    
    if output_file is None:
        output_file = os.path.splitext(excel_file)[0] + '.tex'
    
    excel_to_latex(excel_file, output_file, caption=caption, label=label)
    
    return output_file
