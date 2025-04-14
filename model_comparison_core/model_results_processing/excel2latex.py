import pandas as pd
import openpyxl
import re
import os
from openpyxl.utils import get_column_letter


def excel_to_latex(excel_file, output_file=None, sheet_name='Results', caption=None, label=None, 
                  include_hidden_columns=False, include_hidden_rows=False, float_format='.3f',
                  font_size='small', resize_to_textwidth=True, use_booktabs=True,
                  vertical_dividers=True, combine_metrics=False, parameter_columns=None,
                  is_single_dataset=False):
    """
    Convert an Excel file to a LaTeX table, preserving bold formatting and asterisks.
    If 'Data Set' column is present, creates a table with metrics grouped by datasets.
    
    Parameters:
    -----------
    excel_file : str
        Path to the Excel file
    output_file : str, optional
        Path to save the LaTeX code. If None, returns the LaTeX code as a string.
    sheet_name : str, optional
        Name of the sheet to convert (default: 'Results')
    caption : str, optional
        Caption for the LaTeX table
    label : str, optional
        Label for the LaTeX table (for cross-referencing)
    include_hidden_columns : bool, optional
        Whether to include hidden columns in the LaTeX table (default: False)
    include_hidden_rows : bool, optional
        Whether to include hidden rows in the LaTeX table (default: False)
    float_format : str, optional
        Format string for floating point numbers (default: '.3f')
    font_size : str, optional
        LaTeX font size command (e.g., 'tiny', 'scriptsize', 'footnotesize', 'small', 'normalsize')
    resize_to_textwidth : bool, optional
        Whether to use \\resizebox to make the table fit the textwidth
    use_booktabs : bool, optional
        Whether to use booktabs package for table rules (toprule, midrule, bottomrule)
    vertical_dividers : bool, optional
        Whether to include vertical dividers in the table
    combine_metrics : bool, optional
        If True, creates a table with metrics as column groups (like in the example)
    parameter_columns : list, optional
        List of column names that define the model parameters (e.g., ['DR', 'Geo', 'Num Dim'])
        If None, will look for a 'Model' column or try to detect parameter columns
    is_single_dataset : bool, optional
        If True, indicates this file contains only one dataset, so don't show dataset names in headers
        
    Returns:
    --------
    str or None
        If output_file is None, returns the LaTeX code as a string.
        Otherwise, writes the LaTeX code to the output file and returns None.
    """
    print(f"Processing Excel file: {excel_file}")
    
    # Open the workbook to check formatting and hidden columns FIRST
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[sheet_name]
    
    # Get list of hidden columns from Excel directly
    hidden_cols = []
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        # Check if column is hidden either by direct property or by width
        is_hidden = (ws.column_dimensions[col_letter].hidden or 
                    getattr(ws.column_dimensions[col_letter], 'width', None) == 0)
        
        # Also check column header for std and p-value indicators
        header_cell = ws.cell(row=1, column=col_idx)
        header_value = str(header_cell.value) if header_cell.value else ""
        is_stat_column = any(suffix in header_value for suffix in [' Std', ' p', ' Stat', ' Diff'])
        
        if (not include_hidden_columns) and (is_hidden or is_stat_column):
            hidden_cols.append(col_idx - 1)  # Convert to 0-indexed for pandas
            print(f"Marking column {header_value} as hidden")
    
    print(f"Found {len(hidden_cols)} hidden columns")
    
    # Get list of hidden rows from Excel
    hidden_rows = []
    for row_idx in range(2, ws.max_row + 1):  # Skip header row (1-indexed)
        # Check if row is hidden
        is_hidden = ws.row_dimensions[row_idx].hidden
        
        if (not include_hidden_rows) and is_hidden:
            hidden_rows.append(row_idx - 2)  # Convert to 0-indexed for pandas (after accounting for header)
            print(f"Marking row {row_idx} as hidden")
    
    print(f"Found {len(hidden_rows)} hidden rows")
    
    # Read the DataFrame, and immediately filter out hidden columns
    # Preserve 'None' and 'NA' as strings instead of converting to NaN
    df = pd.read_excel(
        excel_file, 
        sheet_name=sheet_name, 
        keep_default_na=False,  # Don't convert 'NA' to NaN
        na_values=[''],  # Only treat empty cells as NaN
        dtype=str  # Read all data as strings initially
    )
    
    # Debug: Print the columns in the DataFrame
    print("Original columns:", df.columns.tolist())
    print(f"Original shape: {df.shape}")
    
    # Check for specific values like None and NA in the DataFrame
    for col in df.columns:
        none_count = df[col].apply(lambda x: str(x).lower() == 'none').sum()
        na_count = df[col].apply(lambda x: str(x).lower() == 'na').sum()
        if none_count > 0 or na_count > 0:
            print(f"Column '{col}' contains: {none_count} 'None' values and {na_count} 'NA' values")
    
    # Initialize latex_code here to avoid UnboundLocalError
    latex_code = []
    
    # Filter out hidden columns from the DataFrame
    if hidden_cols:
        visible_columns = [col for i, col in enumerate(df.columns) if i not in hidden_cols]
        df_visible = df[visible_columns]
        print(f"Filtered DataFrame from {len(df.columns)} columns to {len(df_visible.columns)} columns")
        print("Visible columns:", visible_columns)
    else:
        visible_columns = list(df.columns)
        df_visible = df
        print("No hidden columns found or hidden columns included")
    
    # Filter out hidden rows from the DataFrame
    if hidden_rows:
        df_visible = df_visible.drop(df_visible.index[hidden_rows])
        print(f"Filtered DataFrame from {len(df)} rows to {len(df_visible)} rows after removing hidden rows")
    
    # Detect table structure and parameter columns
    # Check for 'Data Set' column with case-insensitive match
    data_set_columns = [col for col in df_visible.columns if col.lower() == 'data set']
    has_dataset_column = len(data_set_columns) > 0
    dataset_column_name = data_set_columns[0] if has_dataset_column else None
    
    # Check for 'Model' column
    model_columns = [col for col in df_visible.columns if col.lower() == 'model']
    has_model_column = len(model_columns) > 0
    model_column_name = model_columns[0] if has_model_column else None
    
    # Debug dataset detection
    print(f"Dataset column detected: {has_dataset_column}, Column name: {dataset_column_name}")
    print(f"Model column detected: {has_model_column}, Column name: {model_column_name}")
    
    # Normalize parameter column names to match case in the DataFrame
    if parameter_columns:
        # Case-insensitive matching of parameter columns
        normalized_params = []
        for param in parameter_columns:
            # Find the actual column name with matching case
            matches = [col for col in df_visible.columns if col.lower() == param.lower()]
            if matches:
                normalized_params.append(matches[0])
            else:
                print(f"Warning: Parameter column '{param}' not found in DataFrame")
        
        parameter_columns = normalized_params
        print(f"Normalized parameter columns: {parameter_columns}")
    
    # Auto-detect parameter columns if not provided
    if not has_model_column and parameter_columns is None:
        # Potential parameter columns (common ML parameters) - using lowercase for comparison
        potential_params = ['dr', 'geo', 'num dim', 'geometry', 'dimension', 'parameters', 
                          'init', 'initialization', 'order', 'input dim', 'geo params', 'layers', 'bc']
        detected_params = [col for col in df_visible.columns if col.lower() in potential_params]
        if detected_params:
            parameter_columns = detected_params
            print(f"Auto-detected parameter columns: {parameter_columns}")
    
    # Check if we have parameter columns (all exist in the DataFrame)
    has_parameter_columns = parameter_columns is not None and len(parameter_columns) > 0 and all(col in df_visible.columns for col in parameter_columns)
    print(f"Using parameter columns: {has_parameter_columns}, Columns: {parameter_columns}")
    
    # If we have dataset column and either Model column or parameter columns, use dataset-grouped formatting
    if has_dataset_column and (has_model_column or has_parameter_columns):
        print("Found 'Data Set' column, using dataset-grouped formatting")
        datasets = sorted(df_visible[dataset_column_name].unique())
        print(f"Found datasets: {datasets}")
        
        # Create a mapping of full dataset names to short names
        dataset_map = {
            'Bimanual': 'BM',
            'CMU': 'CMU',
            'BM': 'BM',  # Add direct mapping
            'CMU': 'CMU',  # Add direct mapping
            # Add more mappings as needed
        }
        
        # Get the row identifier column(s)
        if has_model_column:
            # Use Model column as row identifier
            row_id_columns = [model_column_name]
            # Get the metrics (excluding 'Model' and 'Data Set' columns)
            metrics = [col for col in df_visible.columns if col not in [model_column_name, dataset_column_name]]
        else:
            # Use parameter columns as row identifiers
            row_id_columns = parameter_columns
            # Get the metrics (excluding parameter columns and 'Data Set')
            metrics = [col for col in df_visible.columns if col not in parameter_columns + [dataset_column_name]]
        
        print(f"Using row identifiers: {row_id_columns}")
        print(f"Found metrics: {metrics}")
        
        # Create a unique row identifier by combining parameter columns
        if not has_model_column and has_parameter_columns:
            # Replace underscores with spaces in parameter columns
            for col in parameter_columns:
                df_visible[col] = df_visible[col].apply(
                    lambda x: str(x).replace('_', ' ') if pd.notnull(x) else x
                )
            
            # Create a Model column by combining parameter columns
            df_visible['Combined_ID'] = df_visible[parameter_columns].apply(
                lambda row: ' + '.join(str(v) for v in row.values if pd.notnull(v)),
                axis=1
            )
            # Store original order of row IDs
            original_row_ids = df_visible['Combined_ID'].unique()
        else:
            # For Model column, replace underscores with spaces
            if model_column_name in df_visible.columns:
                df_visible[model_column_name] = df_visible[model_column_name].apply(
                    lambda x: str(x).replace('_', ' ') if pd.notnull(x) else x
                )
            df_visible['Combined_ID'] = df_visible[model_column_name]
            original_row_ids = df_visible[model_column_name].unique()
        
        print(f"Original row ID order: {original_row_ids}")
        
        # Debug pivot table creation
        try:
            # Create a pivot table for easier access to dataset-specific values
            # Use the Combined_ID as index
            if has_parameter_columns:
                # For parameter columns, we'll keep them separate in the output
                print(f"Creating pivot table with parameters: {parameter_columns}")
                print(f"Dataset column: {dataset_column_name}")
                
                # Create the pivot table
                df_pivot = df_visible.pivot_table(
                    index=parameter_columns, 
                    columns=dataset_column_name,
                    values=metrics,
                    aggfunc='first',  # In case of duplicates, take the first value
                    dropna=False  # Keep missing values
                )
                
                print("Pivot table created successfully")
                print(f"Pivot table shape: {df_pivot.shape}")
                
                # But we need to create a mapping from combined ID to parameter values
                param_mapping = {}
                for _, row in df_visible.drop_duplicates(parameter_columns).iterrows():
                    param_key = tuple(row[col] for col in parameter_columns)
                    param_mapping[param_key] = [row[col] for col in parameter_columns]
                
                # Get unique parameter combinations in original order
                unique_params = []
                for row_id in original_row_ids:
                    for idx, row in df_visible.iterrows():
                        if row['Combined_ID'] == row_id:
                            param_key = tuple(row[col] for col in parameter_columns)
                            if param_key not in unique_params:
                                unique_params.append(param_key)
                                print(f"Added parameter key: {param_key}")
                            break
                
                # Reindex to preserve original order
                if unique_params:
                    try:
                        print(f"Reindexing pivot table with {len(unique_params)} unique parameters")
                        df_pivot = df_pivot.reindex(unique_params)
                    except Exception as e:
                        print(f"Error reindexing pivot table: {str(e)}")
                        # If reindexing fails, keep the pivot table as is
                        print("Using pivot table without reindexing")
            else:
                # For Model column, we can use it directly
                df_pivot = df_visible.pivot(
                    index='Combined_ID', 
                    columns=dataset_column_name
                )
                # Ensure NaN values are preserved
                df_pivot = df_pivot.fillna(value=pd.NA)
                # Reorder the index to match original order
                df_pivot = df_pivot.reindex(original_row_ids)
        
        except Exception as e:
            print(f"Error creating pivot table: {str(e)}")
            print("Falling back to standard table format")
            # Fall back to standard table format
            has_dataset_column = False
        
        if has_dataset_column:  # Continue only if pivot table was created successfully
            # Create a map of bold cells based on the visible columns
            bold_cells = {}
            for row_idx in range(2, ws.max_row + 1):  # Skip header row, use Excel 1-indexed
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.font.bold:
                        # We'll store the original cell value to match with our pivot table
                        cell_value = str(cell.value) if cell.value else ""
                        bold_cells[(row_idx - 2, col_idx - 1)] = cell_value
            
            # Start building the LaTeX table
            latex_code = []
            
            # Add preamble
            latex_code.append("\\begin{table}[ht]")
            
            if caption:
                latex_code.append(f"\\caption{{{caption}}}")
            
            if label:
                latex_code.append(f"\\label{{{label}}}")
            
            latex_code.append("\\centering")
            
            # Add font size
            if font_size:
                latex_code.append(f"\\{font_size}")
            
            # Add resizebox if requested
            if resize_to_textwidth:
                latex_code.append("\\resizebox{\\textwidth}{!}{%")
            
            # Create column format string
            n_datasets = len(datasets)
            if has_parameter_columns:
                # Format for parameter columns (left-aligned)
                col_format = 'l' * len(parameter_columns)
            else:
                # Format for Model column (left-aligned) 
                col_format = 'l'
                
            # Format for metric columns (center-aligned for dataset headers, right-aligned for values)
            for _ in metrics:
                if vertical_dividers:
                    col_format += '|' + 'c' * n_datasets
                else:
                    col_format += 'c' * n_datasets
                    
            if vertical_dividers:
                col_format = '|' + col_format + '|'
            
            # Begin tabular environment
            latex_code.append(f"\\begin{{tabular}}{{{col_format}}}")
            
            # Add toprule if using booktabs
            if use_booktabs:
                latex_code.append("\\toprule")
            else:
                latex_code.append("\\hline")
            
            # Add multicolumn headers for metrics
            header1_parts = []
            header2_parts = []
            
            # Add parameter column headers or Model header
            if has_parameter_columns:
                # First header row has empty cells for parameter columns
                header1_parts.extend(['' for _ in parameter_columns])
                # Second header row has parameter column names
                header2_parts.extend([f"\\textbf{{{col}}}" for col in parameter_columns])
            else:
                # For Model column
                header1_parts.append('')
                header2_parts.append('\\textbf{Model}')
            
            # Add metric column headers
            for metric in metrics:
                if vertical_dividers:
                    header1_parts.append(f"\\multicolumn{{{n_datasets}}}{{c|}}{{{f'\\textbf{{{metric}}}'}}}") 
                else:
                    header1_parts.append(f"\\multicolumn{{{n_datasets}}}{{c}}{{{f'\\textbf{{{metric}}}'}}}") 
                
                # Check if this is a single dataset table and we should suppress dataset names in headers
                if is_single_dataset:
                    # For single dataset table, just add an empty column header
                    header2_parts.extend(['' for _ in datasets])
                else:
                    # For multi-dataset table, add each dataset name
                    header2_parts.extend([f"\\textbf{{{dataset_map.get(ds, ds)}}}" for ds in datasets])
            
            latex_code.append(" & ".join(header1_parts) + " \\\\")
            
            # Always include the second header row even in single dataset mode
            # because it contains parameter column names (Geo, Dims, DR)
            latex_code.append(" & ".join(header2_parts) + " \\\\")
            
            # Add midrule if using booktabs
            if use_booktabs:
                latex_code.append("\\midrule")
            else:
                latex_code.append("\\hline")
            
            # Add data rows
            if has_parameter_columns:
                # For parameter columns table
                for param_key in df_pivot.index:
                    # Get parameter values for this row
                    row_parts = []
                    # Add parameter values
                    for value in param_key:
                        if pd.isna(value) and not (isinstance(value, str) and value.lower() in ['none', 'na']):
                            # Only replace NaN values (not 'None' or 'NA' strings) with ---
                            row_parts.append("---")
                        else:
                            # Check if value is a string representation of None or NA
                            str_value = str(value)
                            if str_value.lower() in ['none', 'na']:
                                # Ensure these values are preserved exactly as they appear in Excel
                                row_parts.append(str_value)
                            else:
                                row_parts.append(str_value)
                    
                    # Add metric values for each dataset
                    for metric in metrics:
                        for dataset in datasets:
                            try:
                                # Access the value using multi-level indexing
                                value = df_pivot.loc[param_key, (metric, dataset)]
                                if pd.isna(value) and not (isinstance(value, str) and value.lower() in ['none', 'na']):
                                    cell_text = "---"
                                else:
                                    if isinstance(value, (int, float)):
                                        cell_text = f"{value:{float_format}}"
                                    else:
                                        # Preserve string values as they are, including 'None' and 'NA'
                                        str_value = str(value)
                                        if str_value.lower() in ['none', 'na']:
                                            # Ensure these values are preserved exactly as they appear in Excel
                                            cell_text = str_value
                                        else:
                                            cell_text = str_value
                                    
                                    # Check for asterisks and convert to superscript
                                    if isinstance(cell_text, str) and cell_text.endswith('*'):
                                        cell_text = cell_text[:-1] + '$^{*}$'
                                    
                                    # Check if this value was bold in the original Excel
                                    if any(str(value) in str(v) for v in bold_cells.values()):
                                        cell_text = f"\\textbf{{{cell_text}}}"
                            except (KeyError, TypeError, ValueError) as e:
                                print(f"Error accessing data for {param_key}, {metric}, {dataset}: {str(e)}")
                                cell_text = "---"
                            
                            row_parts.append(cell_text)
                    
                    latex_code.append(" & ".join(row_parts) + " \\\\")
            else:
                # For Model column table
                for model in df_pivot.index:
                    row_parts = [f"{model}"]  # Model name
                    for metric in metrics:
                        for dataset in datasets:
                            try:
                                value = df_pivot.loc[model, (metric, dataset)]
                                if pd.isna(value) and not (isinstance(value, str) and value.lower() in ['none', 'na']):
                                    cell_text = "---"
                                else:
                                    if isinstance(value, (int, float)):
                                        cell_text = f"{value:{float_format}}"
                                    else:
                                        # Preserve string values as they are, including 'None' and 'NA'
                                        str_value = str(value)
                                        if str_value.lower() in ['none', 'na']:
                                            # Ensure these values are preserved exactly as they appear in Excel
                                            cell_text = str_value
                                        else:
                                            cell_text = str_value
                                    
                                    # Check for asterisks and convert to superscript
                                    if isinstance(cell_text, str) and cell_text.endswith('*'):
                                        cell_text = cell_text[:-1] + '$^{*}$'
                                    
                                    # Check if this value was bold in the original Excel
                                    if any(str(value) in str(v) for v in bold_cells.values()):
                                        cell_text = f"\\textbf{{{cell_text}}}"
                            except (KeyError, TypeError, ValueError) as e:
                                print(f"Error accessing data for {model}, {metric}, {dataset}: {str(e)}")
                                cell_text = "---"
                            
                            row_parts.append(cell_text)
                    
                    latex_code.append(" & ".join(row_parts) + " \\\\")
            
            # Close the table
            if use_booktabs:
                latex_code.append("\\bottomrule")
            else:
                latex_code.append("\\hline")
            
            latex_code.append("\\end{tabular}")
            
            # Close resizebox if used
            if resize_to_textwidth:
                latex_code.append("}")
            
            latex_code.append("\\end{table}")
        
        else:
            # Fall back to standard table format (the following is the existing code)
            # Create a map of bold cells based on the visible columns
            bold_cells = {}
            for row_idx in range(2, ws.max_row + 1):  # Skip header row, use Excel 1-indexed
                for visible_col_idx, column in enumerate(visible_columns):
                    # Find the original column index in the Excel file
                    excel_col_idx = list(df.columns).index(column) + 1  # 1-indexed for Excel
                    
                    # Check if the cell has bold formatting
                    cell = ws.cell(row=row_idx, column=excel_col_idx)
                    if cell.font.bold:
                        # Convert to 0-indexed for DataFrame
                        bold_cells[(row_idx - 2, visible_col_idx)] = True
            
            print(f"Found {len(bold_cells)} bold cells in Excel")
            
            # Replace underscores with spaces in Model column
            model_col_idx = 0  # Assuming Model is the first column
            if 'Model' in df_visible.columns or df_visible.columns[0] == 'Model':
                print("Replacing underscores with spaces in Model column")
                # Create a copy to avoid SettingWithCopyWarning
                df_visible = df_visible.copy()
                df_visible.iloc[:, model_col_idx] = df_visible.iloc[:, model_col_idx].apply(
                    lambda x: str(x).replace('_', ' ') if pd.notnull(x) else x
                )
            
            # Start building the LaTeX table
            latex_code = []
            
            # Add preamble
            latex_code.append("\\begin{table}[ht]")
            
            if caption:
                latex_code.append(f"\\caption{{{caption}}}")
            
            if label:
                latex_code.append(f"\\label{{{label}}}")
            
            latex_code.append("\\centering")
            
            # Add font size
            if font_size:
                latex_code.append(f"\\{font_size}")
            
            # Add resizebox if requested
            if resize_to_textwidth:
                latex_code.append("\\resizebox{\\textwidth}{!}{%")
            
            # Determine column format
            if combine_metrics:
                # This is for a specialized table with metrics as column groups
                # We'll need to implement this separately
                pass
            else:
                # Column format: l for the first column (model names), c for parameters, r for numeric data
                col_formats = ['l']  # First column (Model) is left-aligned
                
                for col in visible_columns[1:]:
                    # For parameter columns, use centered alignment
                    if col == 'Model' or col in ['Init', 'Input Dim', 'Geo Params', 'Order', 'Geometry'] or 'Params' in col:
                        col_formats.append('c')
                    else:
                        # For numeric columns, use right alignment
                        col_formats.append('r')
                
                # Add vertical dividers if requested
                if vertical_dividers:
                    col_format_str = '|'.join(col_formats)
                    col_format_str = '|' + col_format_str + '|'
                else:
                    col_format_str = ''.join(col_formats)
            
            # Begin tabular environment
            latex_code.append(f"\\begin{{tabular}}{{{col_format_str}}}")
            
            # Add toprule if using booktabs
            if use_booktabs:
                latex_code.append("\\toprule")
            else:
                latex_code.append("\\hline")
            
            # Add header row with bold formatting
            header_row = " & ".join([f"\\textbf{{{col}}}" for col in visible_columns])
            latex_code.append(f"{header_row} \\\\")
            
            # Add midrule if using booktabs
            if use_booktabs:
                latex_code.append("\\midrule")
            else:
                latex_code.append("\\hline")
            
            # Add data rows
            for row_idx, row in df_visible.iterrows():
                row_data = []
                
                for col_idx, (col_name, value) in enumerate(row.items()):
                    # Format the cell value
                    if pd.isna(value) and not (isinstance(value, str) and value.lower() in ['none', 'na']):
                        cell_text = "---"
                    elif isinstance(value, (int, float)):
                        # Format number with appropriate precision
                        if col_name.startswith(('F1', 'Distance', 'Dampening', 'LDJ')) or col_name == 'Cumulative':
                            cell_text = f"{value:{float_format}}"
                        else:
                            cell_text = str(value)
                    else:
                        # Preserve string values as they are, including 'None' and 'NA'
                        str_value = str(value)
                        if str_value.lower() in ['none', 'na']:
                            # Ensure these values are preserved exactly as they appear in Excel
                            cell_text = str_value
                        else:
                            cell_text = str_value
                    
                    # Check for asterisks and convert to superscript
                    if isinstance(cell_text, str) and cell_text.endswith('*'):
                        cell_text = cell_text[:-1] + '$^{*}$'
                    
                    # Apply bold formatting if needed
                    if (row_idx, col_idx) in bold_cells:
                        cell_text = f"\\textbf{{{cell_text}}}"
                    
                    row_data.append(cell_text)
                
                # Join the row data and add to latex code
                latex_code.append(" & ".join(row_data) + " \\\\")
            
            # Close the table
            if use_booktabs:
                latex_code.append("\\bottomrule")
            else:
                latex_code.append("\\hline")
            
            latex_code.append("\\end{tabular}")
            
            # Close resizebox if used
            if resize_to_textwidth:
                latex_code.append("}")
            
            latex_code.append("\\end{table}")
    
    # Compile the full LaTeX code
    full_latex = "\n".join(latex_code)
    
    # Add package requirements comment
    package_comment = "% Requires: "
    packages = []
    if use_booktabs:
        packages.append("\\usepackage{booktabs}")
    package_comment += ", ".join(packages)
    full_latex = package_comment + "\n\n" + full_latex if packages else full_latex
    
    # Save to file or return as string
    if output_file:
        with open(output_file, 'w') as f:
            f.write(full_latex)
        print(f"LaTeX table saved to {output_file}")
        return output_file  # Return the output file path for convenience
    else:
        return full_latex


def excel_to_latex_by_dataset(excel_file, output_file=None, sheet_name='Results', caption=None, label=None, 
                           include_hidden_columns=False, include_hidden_rows=False, float_format='.3f',
                           font_size='small', resize_to_textwidth=True, use_booktabs=True,
                           vertical_dividers=True, parameter_columns=None):
    """
    Convert an Excel file to multiple LaTeX tables, one for each dataset.
    
    Parameters:
    -----------
    excel_file : str
        Path to the Excel file
    output_file : str, optional
        Base path to save the LaTeX code. Dataset name will be appended before the extension.
    sheet_name : str, optional
        Name of the sheet to convert (default: 'Results')
    caption : str, optional
        Caption for the LaTeX table. Dataset name will be appended.
    label : str, optional
        Label for the LaTeX table. Dataset name will be appended.
    include_hidden_columns : bool, optional
        Whether to include hidden columns in the LaTeX table (default: False)
    include_hidden_rows : bool, optional
        Whether to include hidden rows in the LaTeX table (default: False)
    float_format : str, optional
        Format string for floating point numbers (default: '.3f')
    font_size : str, optional
        LaTeX font size command (e.g., 'tiny', 'scriptsize', 'footnotesize', 'small', 'normalsize')
    resize_to_textwidth : bool, optional
        Whether to use \\resizebox to make the table fit the textwidth
    use_booktabs : bool, optional
        Whether to use booktabs package for table rules (toprule, midrule, bottomrule)
    vertical_dividers : bool, optional
        Whether to include vertical dividers in the table
    parameter_columns : list, optional
        List of column names that define the model parameters
        
    Returns:
    --------
    list
        List of output file paths created
    """
    # Read the Excel file
    df = pd.read_excel(
        excel_file, 
        sheet_name=sheet_name, 
        keep_default_na=False,  # Don't convert 'NA' to NaN
        na_values=[''],  # Only treat empty cells as NaN
        dtype=str  # Read all data as strings initially
    )
    
    # Look for a Data Set column (case-insensitive)
    data_set_columns = [col for col in df.columns if col.lower() == 'data set']
    if not data_set_columns:
        raise ValueError("No 'Data Set' column found in the Excel file. Cannot split by dataset.")
    
    dataset_column = data_set_columns[0]
    datasets = df[dataset_column].unique().tolist()
    
    print(f"Found datasets: {datasets}")
    output_files = []
    
    # Open the workbook directly to read bold formatting
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[sheet_name]
    
    # Create a temporary directory to store intermediate Excel files
    import tempfile
    import shutil
    temp_dir = tempfile.mkdtemp()
    
    try:
        for dataset in datasets:
            # Filter data for this dataset
            df_dataset = df[df[dataset_column] == dataset].copy()
            
            # Store which rows in the original Excel correspond to this dataset
            # This will help us map bold formatting correctly
            orig_row_indices = []
            for i, row in df.iterrows():
                if row[dataset_column] == dataset:
                    orig_row_indices.append(i + 2)  # +2 because Excel is 1-indexed and has a header row
            
            # Keep the Data Set column but set all values to the same dataset
            # This avoids the UnboundLocalError in excel_to_latex
            df_dataset[dataset_column] = dataset
            
            # Create a temporary Excel file for this dataset
            temp_excel = os.path.join(temp_dir, f"temp_{dataset}.xlsx")
            df_dataset.to_excel(temp_excel, sheet_name=sheet_name, index=False)
            
            # Open the newly created Excel file to copy bold formatting
            temp_wb = openpyxl.load_workbook(temp_excel)
            temp_ws = temp_wb[sheet_name]
            
            # Copy bold formatting from original Excel file to the temporary one
            for new_row_idx, orig_row_idx in enumerate(orig_row_indices, start=2):  # Start at 2 to skip header
                for col_idx in range(1, ws.max_column + 1):
                    orig_cell = ws.cell(row=orig_row_idx, column=col_idx)
                    if orig_cell.font.bold:
                        if col_idx <= temp_ws.max_column:  # Make sure column exists in temp file
                            temp_cell = temp_ws.cell(row=new_row_idx, column=col_idx)
                            temp_cell.font = openpyxl.styles.Font(bold=True)
            
            # Save the temporary Excel with bold formatting
            temp_wb.save(temp_excel)
            
            # Generate dataset-specific output path
            if output_file:
                base_name, ext = os.path.splitext(output_file)
                dataset_output = f"{base_name}_{dataset}{ext}"
            else:
                base_name, ext = os.path.splitext(excel_file)
                dataset_output = f"{base_name}_{dataset}.tex"
            
            # Generate dataset-specific caption and label
            # Don't append dataset name since each file will only contain that dataset
            dataset_caption = caption if caption else f"Results for {dataset}"
            dataset_label = f"{label}_{dataset.lower()}" if label else f"tab:results-{dataset.lower()}"
            
            # Convert to LaTeX with is_single_dataset flag to avoid showing dataset in headers
            excel_to_latex(
                temp_excel,
                dataset_output,
                sheet_name=sheet_name,
                caption=dataset_caption,
                label=dataset_label,
                include_hidden_columns=include_hidden_columns,
                include_hidden_rows=include_hidden_rows,
                float_format=float_format,
                font_size=font_size,
                resize_to_textwidth=resize_to_textwidth,
                use_booktabs=use_booktabs,
                vertical_dividers=vertical_dividers,
                parameter_columns=parameter_columns,
                is_single_dataset=True  # Tell the function this is a single dataset file
            )
            
            output_files.append(dataset_output)
            print(f"Created LaTeX table for {dataset} dataset: {dataset_output}")
    
    finally:
        # Clean up temporary directory
        shutil.rmtree(temp_dir)
    
    return output_files


def convert_latest_excel(directory, output_file=None, caption=None, label=None,
                     font_size='small', resize_to_textwidth=True, use_booktabs=True,
                     vertical_dividers=True, float_format='.3f', include_hidden_columns=False,
                     include_hidden_rows=False, parameter_columns=None):
    """
    Find the most recent Excel file in the specified directory and convert it to LaTeX.
    
    Parameters:
    -----------
    directory : str
        Directory containing Excel files
    output_file : str, optional
        Path to save the LaTeX code. If None, saves with the same name as the Excel file but with .tex extension.
    caption : str, optional
        Caption for the LaTeX table
    label : str, optional
        Label for the LaTeX table (for cross-referencing)
    font_size : str, optional
        LaTeX font size command (e.g., 'tiny', 'scriptsize', 'footnotesize', 'small', 'normalsize')
    resize_to_textwidth : bool, optional
        Whether to use \\resizebox to make the table fit the textwidth
    use_booktabs : bool, optional
        Whether to use booktabs package for table rules (toprule, midrule, bottomrule)
    vertical_dividers : bool, optional
        Whether to include vertical dividers in the table
    float_format : str, optional
        Format string for floating point numbers (default: '.3f')
    include_hidden_columns : bool, optional
        Whether to include hidden columns in the LaTeX table (default: False)
    include_hidden_rows : bool, optional
        Whether to include hidden rows in the LaTeX table (default: False)
    parameter_columns : list, optional
        List of column names that define the model parameters (e.g., ['DR', 'Geo', 'Num Dim'])
        
    Returns:
    --------
    str
        Path to the output LaTeX file
    """
    # Find all Excel files in the directory
    excel_files = []
    for file in os.listdir(directory):
        if file.endswith('.xlsx') and 'combined_mccv_analysis' in file:
            excel_files.append(os.path.join(directory, file))
    
    if not excel_files:
        raise FileNotFoundError(f"No Excel files found in {directory}")
    
    # Sort by modification time (newest first)
    latest_excel = max(excel_files, key=os.path.getmtime)
    print(f"Converting most recent Excel file: {latest_excel}")
    
    # Generate output file name if not provided
    if output_file is None:
        base_name = os.path.splitext(os.path.basename(latest_excel))[0]
        output_file = os.path.join(directory, f"{base_name}.tex")
    
    # Convert to LaTeX
    return excel_to_latex(
        latest_excel, 
        output_file, 
        caption=caption, 
        label=label, 
        font_size=font_size,
        resize_to_textwidth=resize_to_textwidth,
        use_booktabs=use_booktabs, 
        vertical_dividers=vertical_dividers,
        float_format=float_format,
        include_hidden_columns=include_hidden_columns,
        include_hidden_rows=include_hidden_rows,
        parameter_columns=parameter_columns
    )


def convert_latest_excel_by_dataset(directory, output_file=None, caption=None, label=None,
                               font_size='small', resize_to_textwidth=True, use_booktabs=True,
                               vertical_dividers=True, float_format='.3f', include_hidden_columns=False,
                               include_hidden_rows=False, parameter_columns=None):
    """
    Find the most recent Excel file in the specified directory and convert it to multiple LaTeX tables by dataset.
    
    Parameters:
    -----------
    directory : str
        Directory containing Excel files
    output_file : str, optional
        Base path to save the LaTeX code. Dataset name will be appended.
    caption : str, optional
        Caption for the LaTeX tables. Dataset name will be appended.
    label : str, optional
        Label for the LaTeX tables. Dataset name will be appended.
    font_size : str, optional
        LaTeX font size command
    resize_to_textwidth : bool, optional
        Whether to use \\resizebox to make the table fit the textwidth
    use_booktabs : bool, optional
        Whether to use booktabs package for table rules
    vertical_dividers : bool, optional
        Whether to include vertical dividers in the table
    float_format : str, optional
        Format string for floating point numbers
    include_hidden_columns : bool, optional
        Whether to include hidden columns in the LaTeX table
    include_hidden_rows : bool, optional
        Whether to include hidden rows in the LaTeX table
    parameter_columns : list, optional
        List of column names that define the model parameters
        
    Returns:
    --------
    list
        List of output file paths created
    """
    # Find all Excel files in the directory
    excel_files = []
    for file in os.listdir(directory):
        if file.endswith('.xlsx') and 'combined_mccv_analysis' in file:
            excel_files.append(os.path.join(directory, file))
    
    if not excel_files:
        raise FileNotFoundError(f"No Excel files found in {directory}")
    
    # Sort by modification time (newest first)
    latest_excel = max(excel_files, key=os.path.getmtime)
    print(f"Converting most recent Excel file: {latest_excel}")
    
    # Generate output file name if not provided
    if output_file is None:
        base_name = os.path.splitext(os.path.basename(latest_excel))[0]
        output_file = os.path.join(directory, f"{base_name}.tex")
    
    # Convert to LaTeX by dataset
    return excel_to_latex_by_dataset(
        latest_excel, 
        output_file, 
        caption=caption, 
        label=label, 
        font_size=font_size,
        resize_to_textwidth=resize_to_textwidth,
        use_booktabs=use_booktabs, 
        vertical_dividers=vertical_dividers,
        float_format=float_format,
        include_hidden_columns=include_hidden_columns,
        include_hidden_rows=include_hidden_rows,
        parameter_columns=parameter_columns
    )


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Convert Excel files to LaTeX tables")
    parser.add_argument("--excel", help="Path to Excel file")
    parser.add_argument("--directory", help="Directory containing Excel files (to find most recent)")
    parser.add_argument("--output", help="Output file path")
    parser.add_argument("--caption", help="Table caption")
    parser.add_argument("--label", help="Table label for cross-referencing")
    parser.add_argument("--include-hidden", action="store_true", help="Include hidden columns")
    parser.add_argument("--include-hidden-rows", action="store_true", help="Include hidden rows")
    parser.add_argument("--font-size", default="small", help="LaTeX font size")
    parser.add_argument("--float-format", default=".3f", help="Format for floating point numbers")
    parser.add_argument("--no-resize", action="store_true", help="Don't resize to textwidth")
    parser.add_argument("--no-booktabs", action="store_true", help="Don't use booktabs")
    parser.add_argument("--no-vertical-dividers", action="store_true", help="Don't use vertical dividers")
    parser.add_argument("--parameter-columns", nargs="+", help="List of column names that define model parameters")
    parser.add_argument("--split-by-dataset", action="store_true", help="Create separate tables for each dataset")
    
    args = parser.parse_args()
    
    if args.split_by_dataset:
        if args.excel:
            excel_to_latex_by_dataset(
                args.excel, 
                args.output, 
                caption=args.caption, 
                label=args.label,
                include_hidden_columns=args.include_hidden,
                include_hidden_rows=args.include_hidden_rows,
                font_size=args.font_size,
                float_format=args.float_format,
                resize_to_textwidth=not args.no_resize,
                use_booktabs=not args.no_booktabs,
                vertical_dividers=not args.no_vertical_dividers,
                parameter_columns=args.parameter_columns
            )
        elif args.directory:
            convert_latest_excel_by_dataset(
                args.directory, 
                args.output, 
                caption=args.caption, 
                label=args.label,
                font_size=args.font_size,
                float_format=args.float_format,
                resize_to_textwidth=not args.no_resize,
                use_booktabs=not args.no_booktabs,
                vertical_dividers=not args.no_vertical_dividers,
                include_hidden_columns=args.include_hidden,
                include_hidden_rows=args.include_hidden_rows,
                parameter_columns=args.parameter_columns
            )
        else:
            print("Please provide either --excel or --directory argument")
    else:
        if args.excel:
            excel_to_latex(
                args.excel, 
                args.output, 
                caption=args.caption, 
                label=args.label,
                include_hidden_columns=args.include_hidden,
                include_hidden_rows=args.include_hidden_rows,
                font_size=args.font_size,
                float_format=args.float_format,
                resize_to_textwidth=not args.no_resize,
                use_booktabs=not args.no_booktabs,
                vertical_dividers=not args.no_vertical_dividers
            )
        elif args.directory:
            convert_latest_excel(
                args.directory, 
                args.output, 
                caption=args.caption, 
                label=args.label,
                font_size=args.font_size,
                float_format=args.float_format,
                resize_to_textwidth=not args.no_resize,
                use_booktabs=not args.no_booktabs,
                vertical_dividers=not args.no_vertical_dividers,
                include_hidden_columns=args.include_hidden,
                include_hidden_rows=args.include_hidden_rows,
                parameter_columns=args.parameter_columns
            )
        else:
            print("Please provide either --excel or --directory argument")
