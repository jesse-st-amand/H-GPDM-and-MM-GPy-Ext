from MCCV_analysis_grouping_base_formal import run_mccv_analysis, print_raw_data, save_results_to_csv
from MCCV_analysis_control_param_func_formal import find_control_and_test_paths
import os
import glob
import pandas as pd
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


param_dir = 'MCCV_ICANN'
save_param_dir = 'intra_model'
test_type = 'ttest'
sub_dir = 'geometries_selected_ICANN'  
# Change from single dataset to a list of datasets
data_sets = ['BM','CMU']  # Example datasets - update as needed
name = 'geos_selected'
model = 'GPDMM'

control_params = {'model': 'best per score',}

dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Lists to collect results for each dataset
all_results = []
all_paths = []  # Store paths to the Excel files
all_bolded_cells = {}  # Store information about which cells were bolded in each dataset

# Process each dataset
for data_set in data_sets:
    print(f"\n\n======= Processing dataset: {data_set} =======")
    
    # Create path for this dataset
    main_dir = dir+r"\output_repository\model_summaries\\" + param_dir + r"\\" + sub_dir + r"\\" + name+"_"+data_set
    print(f"Main directory: {main_dir}")
    
    # Find control and test directories for this dataset
    control_path, test_paths = find_control_and_test_paths(main_dir, control_params)
    
    # Check if we found valid paths
    if not control_path or not test_paths:
        print(f"Error: Could not find valid control path or test paths for dataset {data_set}!")
        continue  # Skip this dataset and move to the next one
    
    # Run the analysis
    print(f"\nRunning MCCV analysis for {data_set}...")
    results = run_mccv_analysis(control_path, test_paths, test_type=test_type)
    all_results.append(results)
    
    # Save individual dataset results
    print(f"\nSaving individual results for {data_set}...")
    save_path = dir+r"\output_repository\\MCCV\\" + save_param_dir + "\\" + sub_dir + r"\\" + name+"_"+data_set + "\\" + test_type + "\\"
    if not os.path.exists(save_path):
        os.makedirs(save_path)
    
    # Save results and get path to the generated Excel file
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = save_results_to_csv(results, save_path)
    
    if excel_path:
        all_paths.append((data_set, excel_path))
        print(f"Analysis for {data_set} complete. Results saved to {excel_path}")
    else:
        print(f"Warning: No results generated for dataset {data_set}")

# Combine results from all datasets if we have more than one
if len(all_paths) > 1:
    print("\n\n======= Creating combined dataset results =======")
    
    # Create combined filename with all dataset names
    datasets_string = "_".join(data_sets)
    combined_save_path = dir+r"\output_repository\\MCCV\\" + save_param_dir + "\\" + sub_dir + r"\\" + name+"_"+datasets_string + "\\" + test_type + "\\"
    if not os.path.exists(combined_save_path):
        os.makedirs(combined_save_path)
    
    # Load each Excel file, add dataset column, and track bolded cells
    all_dfs = []
    for data_set, excel_path in all_paths:
        try:
            # Load the workbook to get formatting information
            wb = load_workbook(excel_path)
            ws = wb.active
            
            # Track which cells are bolded in each column
            bolded_cells = {}
            
            # First, get column indices from headers
            header_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
            col_indices = {}
            for idx, cell in enumerate(header_row, start=1):
                if cell.value:
                    col_indices[cell.value] = idx
            
            # Now check each column of interest for bold cells
            metrics = ['F1', 'Distance', 'Dampening', 'LDJ', 'Cumulative']
            for metric in metrics:
                if metric in col_indices:
                    col_idx = col_indices[metric]
                    bolded_cells[metric] = []
                    
                    # Check each row in this column
                    for row_idx in range(2, ws.max_row + 1):  # Skip header
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.font.bold:
                            # Get the value from this cell to help match it later
                            bolded_cells[metric].append(cell.value)
            
            # Load the data as DataFrame
            df = pd.read_excel(excel_path, keep_default_na=False, na_values=[''])
            
            # Add Data Set column
            df['Data Set'] = data_set
            
            # Ensure 'None' values are preserved
            for col in df.columns:
                # Convert NaN to 'None' string
                df[col] = df[col].apply(lambda x: 'None' if pd.isna(x) or x == '' else x)
            
            # Store DataFrame and bolded cell info
            all_dfs.append(df)
            all_bolded_cells[data_set] = bolded_cells
            
            print(f"Loaded {data_set} data with formatting information")
            
        except Exception as e:
            print(f"Error loading {data_set} Excel file: {str(e)}")
    
    if all_dfs:
        # Combine all DataFrames
        combined_df = pd.concat(all_dfs, ignore_index=True)
        
        # Move Data Set to be the first column
        cols = combined_df.columns.tolist()
        cols.insert(0, cols.pop(cols.index('Data Set')))
        combined_df = combined_df[cols]
        
        # Ensure 'None' values are preserved in the combined DataFrame
        for col in combined_df.columns:
            # Replace any empty strings or NaN values with 'None'
            combined_df[col] = combined_df[col].apply(lambda x: 'None' if pd.isna(x) or x == '' else x)
        
        # Save combined results
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        combined_file = os.path.join(combined_save_path, f"combined_mccv_analysis_{datasets_string}_{timestamp}.xlsx")
        
        # Save to Excel with formatting
        try:
            # Create Excel writer
            with pd.ExcelWriter(combined_file, engine='openpyxl') as writer:
                combined_df.to_excel(writer, index=False, sheet_name='Results')
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Results']
                
                # Format headers - make them bold
                header_font = Font(bold=True)
                for col_idx, col_name in enumerate(combined_df.columns):
                    col_letter = get_column_letter(col_idx + 1)
                    cell = worksheet[f"{col_letter}1"]
                    cell.font = header_font
                    
                    # Hide model column, standard deviation, p-value, statistics, and difference columns
                    if col_name == 'Model' or any(suffix in col_name for suffix in [' Std', ' p', ' Stat', ' Diff']):
                        worksheet.column_dimensions[col_letter].hidden = True
                
                # Set column width for visible columns
                for col_idx, col_name in enumerate(combined_df.columns):
                    col_letter = get_column_letter(col_idx + 1)
                    # Only adjust width for visible columns
                    if col_name != 'Model' and not any(suffix in col_name for suffix in [' Std', ' p', ' Stat', ' Diff']):
                        max_length = 0
                        for row_idx in range(1, len(combined_df) + 2):  # Include header and all rows
                            cell = worksheet[f"{col_letter}{row_idx}"]
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        adjusted_width = max(max_length + 2, 10)  # Add padding and minimum width
                        worksheet.column_dimensions[col_letter].width = adjusted_width
                
                # Apply bold formatting based on the original datasets
                data_set_col_idx = combined_df.columns.get_loc('Data Set') + 1  # +1 for openpyxl indexing
                
                # Process each metric column to apply bold formatting
                metrics = ['F1', 'Distance', 'Dampening', 'LDJ', 'Cumulative']
                for metric in metrics:
                    if metric in combined_df.columns:
                        metric_col_idx = combined_df.columns.get_loc(metric) + 1  # +1 for openpyxl indexing
                        
                        # Go through each row
                        for row_idx in range(2, worksheet.max_row + 1):  # Skip header
                            data_set_cell = worksheet.cell(row=row_idx, column=data_set_col_idx)
                            metric_cell = worksheet.cell(row=row_idx, column=metric_col_idx)
                            
                            if data_set_cell.value in all_bolded_cells:
                                # Check if this value was bolded in the original dataset
                                if metric in all_bolded_cells[data_set_cell.value] and metric_cell.value in all_bolded_cells[data_set_cell.value][metric]:
                                    metric_cell.font = Font(bold=True)
            
            print(f"Combined results from all datasets saved to {combined_file}")
            print(f"Bold formatting from individual datasets preserved in the combined table")
            
        except Exception as e:
            print(f"Error saving combined Excel file: {str(e)}")
    else:
        print("Warning: No data available to combine.")

print("\nAll processing complete.")