import os
import glob
import csv
import numpy as np
import pandas as pd
from scipy import stats
import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def find_best_iteration(filepath):
    """First pass to find best iteration based on validation data"""
    with open(filepath, 'r') as file:
        content = file.read()

    parts = content.split('--- ITERATION RESULTS ---')
    second_table = pd.read_csv(io.StringIO(parts[1].strip()))
    
    # Filter for validation rows
    validation_rows = second_table[second_table['label'] == 'validation']
    
    if validation_rows.empty:
        # If no validation data, fall back to all data
        print(f"Warning: No validation data found in {filepath}. Using all data.")
        validation_rows = second_table
    
    # Find the maximum F1 score
    max_f1 = validation_rows['f1'].max()
    
    # Filter rows with maximum F1 score
    max_f1_rows = validation_rows[validation_rows['f1'] == max_f1]
    
    # Among these rows, find the one with minimum score
    best_validation_row = max_f1_rows.loc[max_f1_rows['score'].idxmin()]
    
    return best_validation_row['iteration']


def process_file(filepath):
    """
    Process a file using a two-pass approach:
    1. Find the best iteration based on validation data
    2. Get test metrics at the best iteration
    """
    #print(f"Processing file: {filepath}")

    # First pass: find the best iteration based on validation data
    best_iteration = find_best_iteration(filepath)
    #print(f"Best iteration found: {best_iteration}")
    
    # Second pass: get test metrics at the best iteration
    with open(filepath, 'r') as file:
        content = file.read()

    parts = content.split('--- ITERATION RESULTS ---')
    second_table = pd.read_csv(io.StringIO(parts[1].strip()))
    
    # Filter for test rows
    test_rows = second_table[second_table['label'] == 'test'].copy()
    
    if test_rows.empty:
        # If no test data, fall back to all data
        print(f"Warning: No test data found in {filepath}. Using all data.")
        test_rows = second_table.copy()
    
    # Find the closest iteration to the best one
    test_rows.loc[:, 'iteration_diff'] = abs(test_rows['iteration'] - best_iteration)
    closest_test_row = test_rows.loc[test_rows['iteration_diff'].idxmin()]
    
    #print(f"Using test metrics from iteration {closest_test_row['iteration']}")
    #print(f"Selected test row:\n{closest_test_row[['iteration', 'score', 'f1', 'smoothness', 'avg_freeze']]}")

    result = {
        'score': closest_test_row['score'],
        'f1': closest_test_row['f1'],
        'smoothness': closest_test_row['smoothness'],
        'freeze': closest_test_row['avg_freeze']
    }

    #print(f"Returned result: {result}")
    return result


def get_model_results(path):
    results = {}
    for file in glob.glob(os.path.join(path, '*.csv')):
        sim_num = int(file.split('_')[-1].split('.')[0])
        results[sim_num] = process_file(file)
    return results


def process_directory(path):
    """
    Process a directory containing CSV result files.
    
    Parameters
    ----------
    path : str
        Path to the directory containing CSV files
    
    Returns
    -------
    dict
        Dictionary with simulation numbers as keys and results as values
    """
    results = {}
    csv_files = glob.glob(os.path.join(path, '*.csv'))
    
    if not csv_files:
        print(f"No CSV files found in {path}")
        return {}
    
    print(f"Found {len(csv_files)} CSV files in {path}")
    
    for file in csv_files:
        try:
            # Extract simulation number from filename
            filename = os.path.basename(file)
            if '_' in filename:
                sim_num = int(filename.split('_')[-1].split('.')[0])
            else:
                sim_num = int(filename.split('.')[0])
                
            # Process the file
            results[sim_num] = process_file(file)
            print(f"  Processed {filename}, simulation {sim_num}")
        except Exception as e:
            print(f"  Error processing {file}: {str(e)}")
    
    return results


def apply_fdr_correction(p_values):
    """
    Apply Benjamini-Hochberg procedure for controlling false discovery rate (FDR)
    
    Parameters:
    -----------
    p_values : list
        List of p-values to correct
        
    Returns:
    --------
    list
        List of corrected p-values in the same order as the input
    """
    if not p_values:
        return []
    
    # Create a list of (index, p-value) tuples
    indexed_p_values = list(enumerate(p_values))
    
    # Sort by p-value
    indexed_p_values.sort(key=lambda x: x[1])
    
    # Get the total number of tests
    m = len(p_values)
    
    # Calculate the critical values and create new p-values
    corrected_p_values = [None] * m
    
    # Apply BH procedure
    for rank, (original_index, p_value) in enumerate(indexed_p_values, 1):
        # Calculate BH critical value
        critical_value = (rank / m) * 0.05
        
        # The corrected p-value is p * m / rank
        corrected_p = min(p_value * m / rank, 1.0)
        
        # Store in the order of the original indices
        corrected_p_values[original_index] = corrected_p
    
    # Ensure monotonicity (each corrected p-value >= previous one when sorted by original p-value)
    for i in range(len(indexed_p_values) - 2, -1, -1):
        original_index = indexed_p_values[i][0]
        next_original_index = indexed_p_values[i + 1][0]
        
        if corrected_p_values[original_index] > corrected_p_values[next_original_index]:
            corrected_p_values[next_original_index] = corrected_p_values[original_index]
    
    return corrected_p_values


def perform_statistical_test(control_data, test_data, test_type='ttest'):
    if test_type == 'ttest':
        statistic, p_value = stats.ttest_rel(control_data, test_data)
    elif test_type == 'wilcoxon':
        statistic, p_value = stats.wilcoxon(control_data, test_data)
    else:
        raise ValueError("Invalid test type. Choose 'ttest' or 'wilcoxon'.")

    return statistic, p_value


def run_mccv_analysis(control_path, test_paths, test_type='ttest'):
    """
    Run Monte Carlo Cross-Validation analysis on control and test directories.

    Parameters:
    -----------
    control_path : str
        Path to the control directory with CSV files
        Special value "BEST_PER_SCORE" means use best model per metric as reference
    test_paths : list
        List of paths to test directories with CSV files
    test_type : str
        Statistical test to perform ('ttest' or 'wilcoxon')
        
    Returns:
    --------
    dict
        Dictionary of results with metrics as keys
    """
    # Special case: Best per score mode
    best_per_score_mode = (control_path == "BEST_PER_SCORE")
    
    if best_per_score_mode:
        print("Using best model per metric as the reference")
        all_paths = test_paths
    else:
        # Standard mode - use the specified control path
        all_paths = [control_path] + test_paths
    
    # Collect all model results
    all_model_results = {}
    for path in all_paths:
        model_name = os.path.basename(path)
        print(f"\nProcessing {model_name}...")
        all_model_results[model_name] = get_model_results(path)
    
    # Initialize the results dictionary
    results = {'f1': {}, 'score': {}, 'freeze': {}, 'smoothness': {}}
    
    # Process each metric separately for best per score mode
    if best_per_score_mode:
        for metric in ['f1', 'score', 'freeze', 'smoothness']:
            # First, calculate mean and standard deviation for each model
            metric_stats = {}
            for model_name, model_data in all_model_results.items():
                # Extract values for this metric from all simulations
                values = [sim_data[metric] for sim_num, sim_data in model_data.items() 
                         if metric in sim_data]
                
                if values:
                    metric_stats[model_name] = {
                        'mean': np.mean(values),
                        'std': np.std(values),
                        'values': values
                    }
            
            # Skip if no models have this metric
            if not metric_stats:
                continue
                
            # Find the best model for this metric
            if metric == 'score':  # Lower is better
                best_model = min(metric_stats.keys(), 
                                key=lambda m: metric_stats[m]['mean'])
                print(f"\n===== {metric.upper()} =====")
                print(f"Finding best model for {metric}...")
                print(f"Best model: {best_model} with {metric} = {metric_stats[best_model]['mean']} (lower is better)")
            else:  # Closer to 1 is better
                best_model = min(metric_stats.keys(), 
                               key=lambda m: abs(metric_stats[m]['mean'] - 1))
                print(f"\n===== {metric.upper()} =====")
                print(f"Finding best model for {metric}...")
                print(f"Best model: {best_model} with {metric} = {metric_stats[best_model]['mean']} (closer to 1 is better)")
            
            # Add best model to results
            results[metric][best_model] = {
                'mean': metric_stats[best_model]['mean'],
                'std': metric_stats[best_model]['std']
            }
            
            # Test each model against the best model
            print(f"Testing all models against reference: {best_model}")
            best_model_values = metric_stats[best_model]['values']
            
            # Store uncorrected p-values temporarily
            uncorrected_results = {}
            p_values = []
            model_names = []
            
            for model_name, model_stats in metric_stats.items():
                if model_name == best_model:
                    continue  # Skip comparing the best model to itself
                
                # Get values for the current model
                model_values = model_stats['values']
                
                # Ensure common number of values
                common_len = min(len(best_model_values), len(model_values))
                if common_len < 2:
                    print(f"Skipping {model_name}: Not enough common data points with {best_model}")
                    continue
                
                # Perform statistical test
                statistic, p_value = perform_statistical_test(
                    best_model_values[:common_len], 
                    model_values[:common_len], 
                    test_type
                )
                
                # Store results temporarily
                uncorrected_results[model_name] = {
                    'mean': model_stats['mean'],
                    'std': model_stats['std'],
                    'difference': model_stats['mean'] - metric_stats[best_model]['mean'],
                    'statistic': statistic,
                    'p_value': p_value
                }
                
                # Save p-value for correction
                p_values.append(p_value)
                model_names.append(model_name)
            
            # Apply FDR correction to all p-values for this metric
            if p_values:
                print(f"  Applying FDR correction to {len(p_values)} p-values...")
                corrected_p_values = apply_fdr_correction(p_values)
                
                # Update results with corrected p-values
                for i, model_name in enumerate(model_names):
                    # Store original p-value for reference
                    uncorrected_p = uncorrected_results[model_name]['p_value']
                    corrected_p = corrected_p_values[i]
                    
                    # Update the result with corrected p-value
                    results[metric][model_name] = {
                        'mean': uncorrected_results[model_name]['mean'],
                        'std': uncorrected_results[model_name]['std'],
                        'difference': uncorrected_results[model_name]['difference'],
                        'statistic': uncorrected_results[model_name]['statistic'],
                        'p_value': corrected_p,
                        'p_value_uncorrected': uncorrected_p
                    }
                    
                    is_significant = corrected_p < 0.05
                    print(f"    {metric} p-value: {uncorrected_p:.4f} (corrected: {corrected_p:.4f}) {'(significant)' if is_significant else ''}")
                
    else:
        # Standard mode (one specific control model)
        # Extract control model results
        control_name = os.path.basename(control_path)
        control_results = all_model_results[control_name]
        
        # Store control model metrics in results
        for metric in ['f1', 'score', 'freeze', 'smoothness']:
            # Extract control values for this metric
            control_values = [data[metric] for sim, data in control_results.items() 
                            if metric in data]
            
            if control_values:
                results[metric][control_name] = {
                    'mean': np.mean(control_values),
                    'std': np.std(control_values)
                }
        
        # For each metric, we will collect all p-values before applying FDR correction
        for metric in ['f1', 'score', 'freeze', 'smoothness']:
            uncorrected_results = {}
            p_values = []
            model_names = []

            # Process test models
            for model_path in test_paths:
                model_name = os.path.basename(model_path)
                test_results = all_model_results[model_name]
                
                # Extract values for this metric from both control and test results
                control_data = []
                test_data = []
                
                # Use simulation numbers that are common to both models
                common_sims = set(control_results.keys()) & set(test_results.keys())

                # Skip metrics that aren't present in both models
                if not all(metric in control_results.get(sim, {}) and 
                          metric in test_results.get(sim, {})
                          for sim in common_sims):
                    print(f"Skipping {metric} for {model_name}: not present in all simulations")
                    continue
                
                # Collect data from common simulations
                for sim in common_sims:
                    if metric in control_results[sim] and metric in test_results[sim]:
                        control_data.append(control_results[sim][metric])
                        test_data.append(test_results[sim][metric])
                
                # Skip if not enough data
                if len(control_data) < 2 or len(test_data) < 2:
                    print(f"Skipping {metric} for {model_name}: not enough data points")
                    continue
                
                # Calculate means
                test_mean = np.mean(test_data)
                control_mean = np.mean(control_data)
                
                # Perform statistical test
                statistic, p_value = perform_statistical_test(control_data, test_data, test_type)

                # Store results temporarily
                uncorrected_results[model_name] = {
                    'mean': test_mean,
                    'std': np.std(test_data),
                    'difference': test_mean - control_mean,
                    'statistic': statistic,
                    'p_value': p_value
                }
                
                # Save p-value for correction
                p_values.append(p_value)
                model_names.append(model_name)
            
            # Apply FDR correction to all p-values for this metric
            if p_values:
                print(f"Applying FDR correction to {len(p_values)} p-values for {metric}...")
                corrected_p_values = apply_fdr_correction(p_values)
                
                # Update results with corrected p-values
                for i, model_name in enumerate(model_names):
                    # Store original p-value for reference
                    uncorrected_p = uncorrected_results[model_name]['p_value']
                    corrected_p = corrected_p_values[i]
                    
                    # Update the result with corrected p-value
                    results[metric][model_name] = {
                        'mean': uncorrected_results[model_name]['mean'],
                        'std': uncorrected_results[model_name]['std'],
                        'difference': uncorrected_results[model_name]['difference'],
                        'statistic': uncorrected_results[model_name]['statistic'],
                        'p_value': corrected_p,
                        'p_value_uncorrected': uncorrected_p
                    }
                    
                    is_significant = corrected_p < 0.05
                    print(f"    {metric} p-value: {uncorrected_p:.4f} (corrected: {corrected_p:.4f}) {'(significant)' if is_significant else ''}")

    return results


def save_results_to_csv(results, output_dir):
    """
    Save the analysis results to an Excel file with formatting.
    - Values are rounded to n decimal places
    - Significant values (p < 0.05) are marked with an asterisk (*)
    - Only the single best performing metric in each column is bolded:
      - For F1: highest value is best
      - For Distance: lowest value is best
      - For Dampening: value closest to 1 is best
      - For LDJ: value closest to 1 is best
      
    Returns:
    --------
    str
        Path to the Excel file with the formatted results, or None if processing failed
    """
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    num_dec = 3
    
    # Known list of geometries to detect when they appear in DR column
    known_geometries = [
            'Ellipse', 'Toroid', 'Linear', 'KB', 'MS', 
            'CCs', 'Torus', 'Sphere', 'Fourier', 'Chebyshev', 
            'Legendre', 'Polynomial', 'Cycloid', 'Hypocycloid', 'RSWs', 'SWs', 'Zernike', 'Toroid', 'Chebyshev', 'Laguerre', 'Legendre', 'SHs', 'Lines'
        ]
    # Add lowercase versions for case-insensitive matching
    known_geometries_lower = [geo.lower() for geo in known_geometries]
    
    # Also add original names for backwards compatibility
    known_geometries_original = [
        'Klein Bottle', 'Mobius Strip', 'Cocentric Circles', 
        'Spherical Harmonics', 'Random Sine Wave', 'Random Sine Waves', 'Sine Wave', 'Sine Waves'
    ]
    for geo in known_geometries_original:
        known_geometries.append(geo)
    known_geometries_lower.extend([geo.lower() for geo in known_geometries_original])
    
    # Function to parse the model name into parameter dictionary
    def parse_model_params(model_name):
        # Start with an empty dictionary for the parameters
        params = {}
        param_order = []  # Track the order of parameters
        
        # Remove control suffix if present, before parsing parameters
        if model_name.endswith(';control'):
            model_name = model_name[:-8]  # Remove ';control' suffix
        
        # Return empty params if model name does not contain parameters
        if '-' not in model_name:
            return params, param_order
            
        # Split by semicolon to get parameter pairs
        param_pairs = model_name.split(';')
        
        for pair in param_pairs:
            if '-' in pair:
                # Split by hyphen to get parameter category and value
                parts = pair.split('-', 1)  # Split on first hyphen only
                if len(parts) == 2:
                    category, value = parts
                    
                    # Format the parameter category (replace _ with space, capitalize words)
                    # Using a standardized way to format parameter names to avoid duplication
                    formatted_category = ' '.join(word.capitalize() for word in category.split('_'))
                    
                    # Ensure consistent naming for Geo Params
                    if formatted_category.lower() == "geo params":
                        formatted_category = "Geo Params"
                    
                    # Apply publication-specific renames
                    if formatted_category == "Init":
                        formatted_category = "DR"
                    elif formatted_category == "Geometry":
                        formatted_category = "Geo"
                    elif formatted_category == "Geo Params":
                        formatted_category = "Dims"
                    
                    # Format the parameter value if it's a string (contains alphabetic characters)
                    if any(c.isalpha() for c in value):
                        value = ' '.join(word.capitalize() for word in value.split('_'))
                    
                    # Special formatting for Dims column - handled after all parameters are parsed
                    if formatted_category == "Dims":
                        # Format as integer if it's a number
                        try:
                            if float(value).is_integer():
                                value = str(int(float(value)))
                        except (ValueError, TypeError):
                            pass
                    
                    # Special formatting for DR column
                    if formatted_category == "DR":
                        # First, check if this is actually a geometry that should be moved to Geo
                        is_geometry = False
                        geometry_name = value
                        
                        # Check if the value is in our known geometries list (case-insensitive)
                        for geo in known_geometries:
                            if geo.lower() in value.lower() or value.lower() in geo.lower():
                                is_geometry = True
                                # Remember the original value before other transformations
                                geometry_name = value
                                break
                        
                        # If it's a geometry, abbreviate it and move to Geo later
                        if is_geometry:
                            # Apply immediate abbreviation to ensure it's moved correctly
                            if "spherical harmonics" in value.lower():
                                value = "SHs"
                            elif "random sine wave" in value.lower() or "random sine waves" in value.lower():
                                value = "RSWs"
                            elif "sine wave" in value.lower() or "sine waves" in value.lower():
                                value = "SWs"
                            elif "mobius strip" in value.lower():
                                value = "MS"
                            elif "klein bottle" in value.lower():
                                value = "KB"
                            elif "cocentric circles" in value.lower() or "spiral" in value.lower():
                                value = "CCs"
                        # Otherwise, apply normal DR transformations
                        else:
                            # Apply specific renamings
                            if value == "Pca":
                                value = "PCA"
                            elif value == "Kernel Pca Rbf" or value.startswith("Kernel"):
                                value = "kPCA"
                            elif value == "Umap Euclidean" or value.startswith("Umap E"):
                                value = "UMAP-E"
                            elif value == "Umap Cosine" or value.startswith("Umap C"):
                                value = "UMAP-C"
                            elif value.startswith("Isomap"):
                                value = "Isomap"
                            
                            # Remove "Basis" from string
                            if "Basis" in value:
                                value = value.replace(" Basis", "")
                    
                    # Special formatting for Geo column
                    if formatted_category == "Geo":
                        # Remove "Basis" from string
                        if "Basis" in value:
                            value = value.replace(" Basis", "")
                        
                        # Apply abbreviations for geometry names
                        if "Spherical Harmonics" in value:
                            value = "SHs"
                        elif "Random Sine Wave" in value or "Random Sine Waves" in value:
                            value = "RSWs"
                        elif "Sine Wave" in value or "Sine Waves" in value:
                            value = "SWs"
                        elif "Mobius Strip" in value:
                            value = "MS"
                        elif "Klein Bottle" in value:
                            value = "KB"
                        elif "Cocentric Circles" in value or value == "Spiral":
                            value = "CCs"
                    
                    params[formatted_category] = value
                    param_order.append(formatted_category)
        
        # Handle special cases for geometry parameters with implicit values
        geometry_param_mapping = {
            'Ellipse': '2',
            'Toroid': '3',
            'Linear': '3',
            'KB': '3',
            'Klein Bottle': '3',
            'MS': '2',
            'Mobius Strip': '2',
            'Spiral': '2',
            'CCs': '2',
            'Cocentric Circles': '2',
            'Torus': '3',
            'Sphere': '3',
        }
        
        # Find the geometry and geo params keys if they exist (case-insensitive)
        geometry_key = next((key for key in params.keys() if key.lower() == "geo" or key.lower() == "geometry"), None)
        geo_params_key = next((key for key in params.keys() if key.lower() == "dims" or key.lower() == "geo params"), None)
        dr_key = next((key for key in params.keys() if key.lower() == "dr" or key.lower() == "init"), None)
        
        # Fix case where a geometry is incorrectly classified as a DR technique
        if dr_key and geometry_key:
            dr_value = params[dr_key]
            geo_value = params[geometry_key]
            
            # Check if Geo is None/NA and DR contains a geometry
            if (geo_value.lower() == 'none' or geo_value.lower() == 'na') and dr_value:
                # Check if the DR value is actually a geometry (case-insensitive)
                is_geometry = False
                matched_geo = None
                
                # Try to match against each known geometry
                for geo in known_geometries:
                    if geo.lower() in dr_value.lower() or dr_value.lower() in geo.lower():
                        is_geometry = True
                        matched_geo = geo
                        break
                
                if is_geometry:
                    # Map to abbreviated form if needed
                    if "spherical harmonics" in dr_value.lower():
                        mapped_value = "SHs"
                    elif "random sine wave" in dr_value.lower() or "random sine waves" in dr_value.lower():
                        mapped_value = "RSWs"
                    elif "sine wave" in dr_value.lower() or "sine waves" in dr_value.lower():
                        mapped_value = "SWs"
                    elif "mobius strip" in dr_value.lower():
                        mapped_value = "MS"
                    elif "klein bottle" in dr_value.lower():
                        mapped_value = "KB"
                    elif "cocentric circles" in dr_value.lower() or "spiral" in dr_value.lower():
                        mapped_value = "CCs"
                    else:
                        # If no specific mapping, use the matched geometry or original value
                        mapped_value = matched_geo if matched_geo in ['Ellipse', 'Toroid', 'Linear', 'Torus', 'Sphere', 'Fourier', 'Chebyshev', 'Legendre', 'Polynomial', 'Cycloid', 'Hypocycloid', 'Zernike', 'Laguerre'] else dr_value
                    
                    # Swap the values - DR becomes None, Geo gets the geometry
                    params[geometry_key] = mapped_value
                    params[dr_key] = 'None'
                    print(f"Corrected: Moved '{dr_value}' from DR to Geo column as '{mapped_value}', set DR to 'None'")
        
        # Now handle 'Na' in Dims based on DR and Geo values
        if geo_params_key and params.get(geo_params_key, '').lower() == 'na':
            dr_value = params.get(dr_key, '')
            geo_value = params.get(geometry_key, '')
            
            # Case 1: DR is None and Geo has a value - change 'Na' to 'All'
            if (dr_value.lower() == 'none' or not dr_value) and geo_value and geo_value.lower() != 'none':
                params[geo_params_key] = 'All'
                print(f"Changed Dims from 'Na' to 'All' (geometry occupies all dimensions)")
            # Case 2: DR has a value and Geo is None - keep as 'NA'
            elif dr_value and dr_value.lower() != 'none' and (geo_value.lower() == 'none' or not geo_value):
                params[geo_params_key] = 'NA'
                print(f"Kept Dims as 'NA' (no geometry present)")
        
        # If there's a Geometry parameter and either:
        # 1. No Geo Params key exists, or
        # 2. Geo Params key exists but has a blank value
        if geometry_key:
            geo_value = params[geometry_key]
            
            # Check if geo params is missing or has a blank value
            missing_geo_params = (geo_params_key is None or 
                                  not params[geo_params_key] or 
                                  params[geo_params_key].strip() == 'Na' or
                                  params[geo_params_key].strip() == '')
            
            if missing_geo_params and geo_value.lower() != 'none':
                # Look for a matching geometry to assign implicit value
                for geometry_name, implicit_value in geometry_param_mapping.items():
                    if geometry_name.lower() in geo_value.lower():
                        # If geo params key already exists, update its value
                        if geo_params_key:
                            params[geo_params_key] = implicit_value
                        else:
                            # Otherwise create a new entry with standard formatting
                            params["Dims"] = implicit_value
                            # Add Dims after Geo in the parameter order
                            if geometry_key in param_order:
                                geo_idx = param_order.index(geometry_key)
                                param_order.insert(geo_idx + 1, "Dims")
                            else:
                                param_order.append("Dims")
                        break
        
        return params, param_order

    # Maps metrics to their new column names
    metric_column_names = {
        'f1': 'F1',
        'score': 'Distance',
        'freeze': 'Dampening',
        'smoothness': 'LDJ'
    }
    
    # Get all unique models across all metrics
    all_models = set()
    for metric in results.keys():
        all_models.update(results[metric].keys())
    
    # Extract parameters from all models once
    all_param_categories = set()
    model_params = {}
    param_orders = {}
    all_param_orders = []  # Collect all parameter orders to determine a consistent order
    
    for model in all_models:
        model_basename = os.path.basename(model)
        params, param_order = parse_model_params(model_basename)
        model_params[model] = params
        param_orders[model] = param_order
        all_param_categories.update(params.keys())
        all_param_orders.append(param_order)
    
    # Create a consistent order for all parameters based on their relative positions
    sorted_models = sorted(all_param_orders, key=len, reverse=True)
    reference_order = sorted_models[0] if sorted_models else []
    
    # Add any missing parameters that exist in other models
    for order in all_param_orders:
        for param in order:
            if param not in reference_order:
                reference_order.append(param)
    
    # Use this reference order for all parameter columns
    param_columns = reference_order
    
    # Prepare rows for the combined table
    combined_rows = []
    
    # Check if we're in "best per score" mode - this mode has different control models per metric
    best_per_score_mode = False
    
    # In best_per_score mode, every metric has its own "best" model as control
    # We can detect this by checking if there's no common control across metrics
    control_models = set()
    for metric in results:
        if metric in ['f1', 'score', 'freeze', 'smoothness'] and results[metric]:
            # Find the model that doesn't have 'difference', 'statistic', or 'p_value' keys
            # This is the control/reference model for this metric
            for model, values in results[metric].items():
                if not any(k in values for k in ['difference', 'statistic', 'p_value']):
                    control_models.add(model)
                    break
    
    # If we have multiple control models, we're in best_per_score mode
    best_per_score_mode = len(control_models) > 1
    
    if best_per_score_mode:
        print("Detected 'best per score' mode - each metric has its own best model as reference")
        # In this mode, we don't have a global control_name, each metric has its own
        control_name = None
    else:
        # Standard mode - use the first model in the first metric as control
        control_name = list(results[list(results.keys())[0]].keys())[0]  # Assume the first model in the first metric is the control
    
    # Process each model
    for model in all_models:
        # In best_per_score mode, a model can be control for some metrics but not others
        # In standard mode, is_control is True only for the global control model
        is_control = (not best_per_score_mode) and (model == control_name)
        
        # Initialize row with model and parameters
        row = {
            'Model': os.path.basename(model)
        }
        
        # Add parameter values
        for param in param_columns:
            row[param] = model_params[model].get(param, 'NA')
        
        # Add metric values
        for metric in ['f1', 'score', 'freeze', 'smoothness']:
            if metric in results and model in results[metric]:
                values = results[metric][model]
                
                # Round mean and std to 2 decimal places
                row[metric_column_names[metric]] = round(values['mean'], num_dec)
                row[f"{metric_column_names[metric]} Std"] = round(values['std'], num_dec)
                
                # Determine if this model is the control for this specific metric
                metric_control = False
                if best_per_score_mode:
                    # In best_per_score mode, the control doesn't have difference/statistic/p_value
                    metric_control = not any(k in values for k in ['difference', 'statistic', 'p_value'])
                
                # Add difference, statistic, p-value for non-control models
                # (either not global control in standard mode or not metric control in best_per_score mode)
                if (not is_control and not metric_control) and 'difference' in values:
                    row[f"{metric_column_names[metric]} Diff"] = round(values['difference'], num_dec)
                    row[f"{metric_column_names[metric]} Stat"] = round(values['statistic'], num_dec)
                    # Use the corrected p-value if available
                    if 'p_value_uncorrected' in values:
                        row[f"{metric_column_names[metric]} p"] = values['p_value']  # This is the corrected p-value
                        row[f"{metric_column_names[metric]} p_uncorr"] = values['p_value_uncorrected']
                    else:
                        row[f"{metric_column_names[metric]} p"] = values['p_value']
                else:
                    # This model is a control for this metric
                    row[f"{metric_column_names[metric]} Diff"] = 'NA'
                    row[f"{metric_column_names[metric]} Stat"] = 'NA'
                    row[f"{metric_column_names[metric]} p"] = 'NA'
                    if 'p_value_uncorrected' in next(iter(results.get(metric, {}).values())) if results.get(metric, {}) else {}:
                        row[f"{metric_column_names[metric]} p_uncorr"] = 'NA'
            else:
                # If this metric doesn't have data for this model, use NA
                row[metric_column_names[metric]] = 'NA'
                row[f"{metric_column_names[metric]} Std"] = 'NA'
                if not is_control:
                    row[f"{metric_column_names[metric]} Diff"] = 'NA'
                    row[f"{metric_column_names[metric]} Stat"] = 'NA'
                    row[f"{metric_column_names[metric]} p"] = 'NA'
                    if 'p_value_uncorrected' in next(iter(results.get(metric, {}).values())) if results.get(metric, {}) else {}:
                        row[f"{metric_column_names[metric]} p_uncorr"] = 'NA'
        
        combined_rows.append(row)
    
    # Create DataFrame from combined rows
    df = pd.DataFrame(combined_rows)
    
    # Add the calculated column: Cumulative (Distance*Dampening/F1)
    df['Cumulative'] = df.apply(
        lambda row: (
            round(float(row['Distance']) * float(row['Dampening']) / float(row['F1']), 3)
            if (row['Distance'] != 'NA' and row['Dampening'] != 'NA' and row['F1'] != 'NA' and float(row['F1']) != 0)
            else 'NA'
        ),
        axis=1
    )
    
    # Sort the DataFrame by the calculated column (Cumulative)
    # First convert the column to numeric, with errors='coerce' to handle 'NA' values
    df['Sort_Value'] = pd.to_numeric(df['Cumulative'], errors='coerce')
    # Sort by the numeric column, with NA values at the end
    df = df.sort_values(by='Sort_Value', ascending=True, na_position='last')
    # Drop the temporary sorting column
    df = df.drop(columns=['Sort_Value'])
    
    # Define column order
    base_columns = param_columns  # Remove 'Model' from base columns to hide it
    metric_columns = []
    for metric in ['F1', 'Distance', 'Dampening', 'LDJ']:
        metric_columns.append(metric)
        metric_columns.append(f"{metric} Std")
        # Only add difference columns for the standard deviation, statistics and p-values
        # if there are non-control models
        if len(all_models) > 1:
            metric_columns.append(f"{metric} Diff")
            metric_columns.append(f"{metric} Stat")
            metric_columns.append(f"{metric} p")
            if 'p_value_uncorrected' in next(iter(results.get('f1', {}).values())) if results.get('f1', {}) else {}:
                metric_columns.append(f"{metric} p_uncorr")
    
    # Add the calculated column at the end
    final_columns = base_columns + metric_columns + ['Cumulative']
    
    # Reorder columns and only include columns that exist
    existing_columns = [col for col in final_columns if col in df.columns]
    df = df[existing_columns]
    
    # Format p-values and apply asterisks to significant values
    p_value_columns = [col for col in df.columns if ' p' in col and 'p_uncorr' not in col]
    for col in p_value_columns:
        metric_col = col.split(' p')[0]
        # Create a new column with asterisks for significant values
        df[metric_col] = df.apply(
            lambda row: str(row[metric_col]) + "*" if pd.notnull(row[col]) and row[col] != 'NA' and float(row[col]) < 0.05 else str(row[metric_col]),
            axis=1
        )
    
    # Save to Excel file
    file_name = f"combined_mccv_analysis_{timestamp}.xlsx"
    file_path = os.path.join(output_dir, file_name)

    # Need to import these libraries for Excel formatting
    try:
        import openpyxl
        from openpyxl.styles import Font
        
        # Create Excel writer with openpyxl engine
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Results')
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Results']
            
            # Format headers - make them bold with light gray background
            header_font = Font(bold=True)
            for col_idx, col_name in enumerate(df.columns):
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                cell = worksheet[f"{col_letter}1"]
                cell.font = header_font
                
                # Hide model column, standard deviation, p-value, statistics, and difference columns
                if col_name == 'Model' or any(suffix in col_name for suffix in [' Std', ' p', ' Stat', ' Diff']):
                    worksheet.column_dimensions[col_letter].hidden = True
            
            # Set column width for visible columns
            for col_idx, col_name in enumerate(df.columns):
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                # Only adjust width for visible columns
                if col_name != 'Model' and not any(suffix in col_name for suffix in [' Std', ' p', ' Stat', ' Diff']):
                    max_length = 0
                    for row_idx in range(1, len(df) + 2):  # Include header and all rows
                        cell = worksheet[f"{col_letter}{row_idx}"]
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = max(max_length + 2, 10)  # Add padding and minimum width
                    worksheet.column_dimensions[col_letter].width = adjusted_width
            
            # Bold only the single best value in each metric column
            metric_names = ['F1', 'Distance', 'Dampening', 'LDJ']
            
            for metric in metric_names:
                # Get index of the metric column
                if metric not in df.columns:
                    continue
                    
                col_idx = df.columns.get_loc(metric)
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                
                # Get all values for this metric
                metric_values = []
                for row_idx, value in enumerate(df[metric]):
                    # Skip NA values
                    if value == 'NA' or pd.isna(value):
                        continue
                        
                    # Check if the value has an asterisk (indicating significance)
                    has_asterisk = False
                    numeric_value = value
                    if isinstance(value, str):
                        if value.endswith('*'):
                            has_asterisk = True
                            numeric_value = value[:-1]  # Remove asterisk for numeric comparison
                    
                    try:
                        # Store as (row_index, numeric_value, has_asterisk)
                        metric_values.append((row_idx, float(numeric_value), has_asterisk))
                    except (ValueError, TypeError):
                        continue
                
                if not metric_values:
                    continue
                
                # Find the best value depending on the metric
                if metric == 'F1':
                    # For F1, higher is better
                    best_value = max(value for _, value, _ in metric_values)
                    best_indices = [idx for idx, value, _ in metric_values if value == best_value]
                elif metric == 'Distance':
                    # For Distance, lower is better
                    best_value = min(value for _, value, _ in metric_values)
                    best_indices = [idx for idx, value, _ in metric_values if value == best_value]
                else:  # LDJ or Dampening
                    # For LDJ and Dampening, closer to 1 is better
                    closest_to_one = min(abs(value - 1) for _, value, _ in metric_values)
                    best_indices = [idx for idx, value, _ in metric_values if abs(value - 1) == closest_to_one]
                
                # Bold only the first (or only) index in best_indices
                if best_indices:
                    idx = best_indices[0]
                    cell = worksheet[f"{col_letter}{idx + 2}"]  # +2 for header row and 1-indexing
                    cell.font = Font(bold=True)
     
        print(f"Combined results saved to {file_path} with formatting")
        print(f"The Model column and statistical columns are hidden in the Excel output")
        print(f"Significant values are marked with asterisks (*)")
        print(f"Only the single best value is bolded in each metric column")
        
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        print("Warning: openpyxl not available. Saving as CSV without formatting.")
        csv_file_path = os.path.join(output_dir, f"combined_mccv_analysis_{timestamp}.csv")
        df.to_csv(csv_file_path, index=False)
        print(f"Combined results saved to {csv_file_path}")
        return None  # Return None since we couldn't create the Excel file
    
    # Return the path to the Excel file
    return file_path

# Debugging function to print raw data
def print_raw_data(results):
    """Print raw data from analysis results for debugging"""
    print("\n=== RAW DATA ===")
    for metric, models in results.items():
        print(f"\nMetric: {metric}")
        for model_name, values in models.items():
            print(f"  Model: {model_name}")
            for key, value in values.items():
                # Format numeric values for readability
                if isinstance(value, (int, float)):
                    print(f"    {key}: {value:.4f}")
                else:
                    print(f"    {key}: {value}")
            print("  ---")


def export_to_latex(excel_file, output_file=None, caption=None, label=None):
    """
    Export the Excel results file to LaTeX format.
    
    Parameters:
    -----------
    excel_file : str
        Path to the Excel file to convert
    output_file : str, optional
        Path where to save the LaTeX file. If None, a file with the same
        name but .tex extension will be created in the same directory.
    caption : str, optional
        Caption for the LaTeX table
    label : str, optional
        Label for the LaTeX table (for cross-referencing)
        
    Returns:
    --------
    str
        Path to the generated LaTeX file
    """
    try:
        from excel2latex import excel_to_latex
    except ImportError:
        from .excel2latex import excel_to_latex
    
    if output_file is None:
        output_file = os.path.splitext(excel_file)[0] + '.tex'
    
    excel_to_latex(excel_file, output_file, caption=caption, label=label)
    
    return output_file
